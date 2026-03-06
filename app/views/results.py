from django.shortcuts import render, redirect, HttpResponseRedirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse
from django.forms import modelformset_factory
from django.db import transaction, IntegrityError
from django.db.models import (
    Avg, Sum, F, Q, Count, Max, Case, When, Value, IntegerField
)
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string

from app.constants import *
from app.models.results import (
    AssessmentType,
    AnnualResult,
    Assessment,
    GradingSystem,
    ReportRemark,
    Result,
    VerificationSample,
    ResultVerificationReport,
    VerificationDiscrepancy,
    VerificationCorrectionLog,
)
from app.forms.results import (
    AssesmentTypeForm, GradingSystemForm, ResultForm, AssessmentForm, BulkAssessmentForm
)
from app.models import *
from app.models.students import Student
from app.models.accounts import StaffAccount
from app.models.classes import ClassSubjectAllocation, AcademicClassStream, AcademicClass, Class, Term
from app.selectors.model_selectors import *
from app.selectors.results import (
    get_grade_and_points, get_current_mode, get_performance_metrics,
    get_grade_from_average, calculate_weighted_subject_averages, get_division,
    get_division_with_override
)
from app.selectors.school_settings import get_current_academic_year
from app.selectors.classes import get_current_term
from app.utils.utils import calculate_grade_and_points
from app.utils.pdf_utils import generate_student_report_pdf
from app.services.results_sampling import (
    ensure_batch_for_assessment,
    submit_batch_for_verification,
    update_sample_mark,
    evaluate_batch_verification,
    reset_batch_to_draft,
    _create_verification_report,
    record_correction,
)

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from collections import Counter, defaultdict
import logging
import tempfile
import io

from xhtml2pdf import pisa

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Font = Border = Side = Alignment = PatternFill = None
    get_column_letter = None

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch, cm
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from django.utils import timezone
from datetime import datetime

logger = logging.getLogger(__name__)

VERIFICATION_ROLE_KEYS = {"admin", "director of studies", "dos"}
MARK_ENTRY_ADMIN_ROLE_KEYS = {
    "admin",
    "head master",
    "headmaster",
    "head teacher",
    "headteacher",
    "director of studies",
    "dos",
}


def _effective_role_name(request):
    active_role = request.session.get("active_role_name")
    if active_role:
        return active_role
    staff_account = getattr(request.user, "staff_account", None)
    if staff_account and getattr(staff_account, "role", None):
        return staff_account.role.name
    return None


def _effective_role_key(request):
    return (_effective_role_name(request) or "").strip().lower()


def _has_verification_access(request):
    if request.user.is_superuser:
        return True
    return _effective_role_key(request) in VERIFICATION_ROLE_KEYS


def _get_staff_account(request):
    staff_account = getattr(request.user, "staff_account", None)
    if staff_account:
        return staff_account
    return StaffAccount.objects.filter(user=request.user).select_related("staff", "role").first()


def _mark_entry_role_key(request, staff_account=None):
    active_role = (request.session.get("active_role_name") or "").strip().lower()
    if active_role:
        return active_role
    if staff_account and getattr(staff_account, "role", None):
        return (staff_account.role.name or "").strip().lower()
    return ""


def _teacher_has_assessment_allocation(staff_member, assessment):
    if not staff_member or not assessment:
        return False
    return ClassSubjectAllocation.objects.filter(
        subject_teacher=staff_member,
        academic_class_stream__academic_class=assessment.academic_class,
        subject=assessment.subject,
    ).exists()


def _mark_entry_denial_reason(request, assessment):
    if request.user.is_superuser:
        return None

    staff_account = _get_staff_account(request)
    role_key = _mark_entry_role_key(request, staff_account=staff_account)

    if role_key in {"class teacher", "class_teacher"}:
        return "Class teachers cannot enter marks. Please contact the subject teacher."

    if role_key == "teacher":
        staff_member = getattr(staff_account, "staff", None)
        if not staff_member:
            return "Your teacher account is not linked to a staff profile."
        if not _teacher_has_assessment_allocation(staff_member, assessment):
            return "You can only enter marks for subjects allocated to you in this class."
        return None

    if role_key in MARK_ENTRY_ADMIN_ROLE_KEYS:
        return None

    return "You do not have permission to enter marks for this assessment."


def _parse_score_or_raise(raw_score, assessment, student_label):
    try:
        score = Decimal((raw_score or "").strip())
    except (ValueError, TypeError, InvalidOperation):
        raise ValueError(f"Invalid score for {student_label}. Use a numeric mark.")

    max_score = Decimal(str(assessment.out_of))
    if score < 0 or score > max_score:
        raise ValueError(
            f"Invalid score for {student_label}. Scores must be between 0 and {assessment.out_of}."
        )
    return score


def _submission_gate_errors(assessment, students):
    errors = []
    if not students:
        errors.append("Cannot submit for verification. This class has no registered students.")
        return errors

    results_qs = Result.objects.filter(
        assessment=assessment,
        student__is_active=True,
    ).select_related("student")
    duplicate_rows = (
        results_qs.values("student_id")
        .annotate(row_count=Count("id"))
        .filter(row_count__gt=1)
    )
    if duplicate_rows.exists():
        duplicate_ids = [item["student_id"] for item in duplicate_rows]
        duplicate_names = list(
            Student.objects.filter(is_active=True, id__in=duplicate_ids)
            .order_by("student_name")
            .values_list("student_name", flat=True)
        )
        preview = ", ".join(duplicate_names[:5]) or "unknown students"
        suffix = "..." if len(duplicate_names) > 5 else ""
        errors.append(
            f"Cannot submit for verification. Duplicate results exist for: {preview}{suffix}."
        )

    expected_ids = {student.id for student in students}
    present_ids = set(results_qs.values_list("student_id", flat=True))
    missing_ids = sorted(expected_ids - present_ids)
    if missing_ids:
        missing_names = [student.student_name for student in students if student.id in set(missing_ids)]
        preview = ", ".join(missing_names[:5]) or "unknown students"
        suffix = "..." if len(missing_names) > 5 else ""
        errors.append(
            f"Cannot submit for verification. Missing marks for {len(missing_ids)} student(s): {preview}{suffix}."
        )

    for result in results_qs:
        try:
            score = Decimal(str(result.score))
        except (ValueError, TypeError, InvalidOperation):
            errors.append(
                f"Cannot submit for verification. Non-numeric mark detected for {result.student}."
            )
            continue
        if score < 0 or score > Decimal(str(assessment.out_of)):
            errors.append(
                f"Cannot submit for verification. Out-of-range mark detected for {result.student}."
            )

    return errors







@login_required
def add_results_view(request, assessment_id=None):
    if not assessment_id:
        return redirect('class_assessment_list')

    # Load the assessment and associated class
    assessment = get_object_or_404(Assessment, id=assessment_id)
    denial_reason = _mark_entry_denial_reason(request, assessment)
    if denial_reason:
        messages.error(request, denial_reason)
        return redirect('list_assessments', class_id=assessment.academic_class_id)
    academic_class = assessment.academic_class

    # Get all registers for the class
    class_registers = ClassRegister.objects.filter(
        academic_class_stream__academic_class=academic_class,
        student__is_active=True,
    ).select_related('student') 

    students = []
    seen_student_ids = set()
    broken_registers = []

    # Safely collect only valid students
    for register in class_registers:
        try:
            student = register.student
            if student.id in seen_student_ids:
                continue
            seen_student_ids.add(student.id)
            students.append(student)
        except Student.DoesNotExist:
            broken_registers.append(register.id)
            continue

    students.sort(key=lambda s: (s.student_name or "", s.reg_no or ""))

    if broken_registers:
        messages.warning(request, f"Some class register entries are (missing student data). Contact admin.")

    # Find existing results and determine who doesn't have one yet
    existing_results = Result.objects.filter(
        assessment=assessment,
        student__is_active=True,
    ).select_related("student", "batch__submitted_by")
    existing_student_ids = {result.student_id for result in existing_results}
    students_without_results = [s for s in students if s.id not in existing_student_ids]

    current_mode = ResultModeSetting.get_mode()
    batch = ensure_batch_for_assessment(assessment)
    all_results_entered = len(students_without_results) == 0 and len(students) > 0

    # Handle form submission
    if request.method == "POST":
        if "save_draft" in request.POST:
            if batch.status not in ["DRAFT", "FLAGGED"]:
                messages.error(request, "Results are locked after submission for verification.")
                return redirect('add_results', assessment_id=assessment.id)

            was_flagged = batch.status == "FLAGGED"
            if was_flagged:
                batch.status = "DRAFT"
                batch.rejection_reason = None
                batch.save(update_fields=["status", "rejection_reason"])
                Result.objects.filter(batch=batch).update(status="DRAFT")

            created_results = []
            updated_count = 0
            correction_count = 0
            result_by_student_id = {item.student_id: item for item in existing_results}
            validated_rows = {}

            for student in students:
                score_raw = (request.POST.get(f"score_{student.id}") or "").strip()
                correction_reason = (request.POST.get(f"remark_{student.id}") or "").strip()
                if not score_raw:
                    continue
                try:
                    score = _parse_score_or_raise(score_raw, assessment, str(student))
                except ValueError as exc:
                    messages.error(request, str(exc))
                    return redirect('add_results', assessment_id=assessment.id)

                existing = result_by_student_id.get(student.id)
                score_changed = bool(existing and existing.score != score)
                if was_flagged and score_changed and not correction_reason:
                    messages.error(
                        request,
                        f"Provide a correction reason for {student} before saving flagged corrections.",
                    )
                    return redirect('add_results', assessment_id=assessment.id)
                validated_rows[student.id] = {"score": score, "reason": correction_reason}

            try:
                with transaction.atomic():
                    for student in students:
                        row_data = validated_rows.get(student.id)
                        if not row_data:
                            continue
                        score = row_data["score"]
                        correction_reason = row_data["reason"]
                        existing = result_by_student_id.get(student.id)
                        if existing:
                            score_changed = existing.score != score
                            if existing.score != score or existing.status != "DRAFT" or existing.batch_id != batch.id:
                                old_score = existing.score
                                existing.score = score
                                existing.status = "DRAFT"
                                existing.batch = batch
                                existing.save(update_fields=["score", "status", "batch"])
                                updated_count += 1
                                if was_flagged and score_changed:
                                    record_correction(
                                        batch=batch,
                                        result=existing,
                                        old_mark=old_score,
                                        new_mark=score,
                                        reason=correction_reason,
                                        user=request.user,
                                    )
                                    VerificationDiscrepancy.objects.filter(batch=batch, result=existing).update(
                                        corrected_mark=score,
                                        action_taken=f"Corrected by {request.user.username}",
                                    )
                                    correction_count += 1
                        else:
                            created_results.append(
                                Result(
                                    assessment=assessment,
                                    student=student,
                                    score=score,
                                    batch=batch,
                                    status="DRAFT",
                                )
                            )

                    if created_results:
                        Result.objects.bulk_create(created_results)
            except IntegrityError:
                messages.error(
                    request,
                    "Could not save draft. Duplicate result rows were detected for one or more students.",
                )
                return redirect('add_results', assessment_id=assessment.id)

            created_count = len(created_results)
            if not created_count and not updated_count:
                messages.info(request, "No changes detected in draft results.")
            else:
                messages.success(
                    request,
                    f"Draft saved successfully. Created {created_count} and updated {updated_count} result(s).",
                )
                if correction_count:
                    messages.info(request, f"Logged {correction_count} correction audit record(s).")

            total_students = len(students)
            total_results = Result.objects.filter(
                assessment=assessment,
                student__is_active=True,
            ).values("student_id").distinct().count()
            if total_students > 0 and total_students == total_results:
                messages.warning(
                    request,
                    "All results are now entered. Submit this batch for verification.",
                )
            return redirect('add_results', assessment_id=assessment.id)

        if "bulk_edit" in request.POST:
            if batch.status not in ["DRAFT", "FLAGGED"]:
                messages.error(request, "Results are locked after submission for verification.")
                return redirect('add_results', assessment_id=assessment.id)

            if batch.status == "FLAGGED":
                batch.rejection_reason = None
                batch.save(update_fields=["rejection_reason"])
                Result.objects.filter(batch=batch).update(status="DRAFT")

            updated_count = 0
            with transaction.atomic():
                for result in existing_results:
                    score_raw = request.POST.get(f"score_{result.student.id}")
                    if score_raw in (None, ""):
                        continue
                    try:
                        score = _parse_score_or_raise(score_raw, assessment, str(result.student))
                    except ValueError as exc:
                        messages.error(request, str(exc))
                        return redirect('add_results', assessment_id=assessment.id)

                    if result.score != score:
                        result.score = score
                        result.save(update_fields=["score"])
                        updated_count += 1

            messages.success(request, f"Updated {updated_count} result(s) successfully!")

            total_students = len(students)
            total_results = Result.objects.filter(
                assessment=assessment,
                student__is_active=True,
            ).count()
            if total_students > 0 and total_students == total_results:
                messages.warning(
                    request,
                    "All results are now entered. Submit this batch for verification.",
                )

            return redirect('add_results', assessment_id=assessment.id)

        if "edit_result" in request.POST:
            result_id = request.POST.get("edit_result")
            result = get_object_or_404(Result, id=result_id, assessment=assessment)
            score_raw = request.POST.get(f'score_{result.student.id}')
            correction_reason = (request.POST.get(f"remark_{result.student.id}") or "").strip()
            try:
                if batch.status not in ["DRAFT", "FLAGGED"]:
                    messages.error(request, "Results are locked after submission for verification.")
                    return redirect('add_results', assessment_id=assessment.id)

                new_score = _parse_score_or_raise(score_raw, assessment, str(result.student))
                score_changed = result.score != new_score
                if batch.status == "FLAGGED":
                    if score_changed and not correction_reason:
                        messages.error(
                            request,
                            f"Provide a correction reason for {result.student} before saving flagged corrections.",
                        )
                        return redirect('add_results', assessment_id=assessment.id)
                    batch.status = "DRAFT"
                    batch.rejection_reason = None
                    batch.save(update_fields=["status", "rejection_reason"])
                    Result.objects.filter(batch=batch).update(status="DRAFT")

                old_score = result.score
                result.score = new_score
                result.save(update_fields=["score"])
                if batch.status == "DRAFT" and score_changed and correction_reason:
                    # keep optional audit in normal draft flow when a reason is supplied
                    record_correction(
                        batch=batch,
                        result=result,
                        old_mark=old_score,
                        new_mark=new_score,
                        reason=correction_reason,
                        user=request.user,
                    )
                if score_changed:
                    VerificationDiscrepancy.objects.filter(batch=batch, result=result).update(
                        corrected_mark=new_score,
                        action_taken=f"Corrected by {request.user.username}",
                    )
                messages.success(request, f"Result for {result.student} updated successfully!")

                total_students = len(students)
                total_results = Result.objects.filter(
                    assessment=assessment,
                    student__is_active=True,
                ).count()
                if total_students > 0 and total_students == total_results:
                    messages.warning(
                        request,
                        "All results are now entered. Submit this batch for verification.",
                    )
            except ValueError as exc:
                messages.error(request, str(exc))
            return redirect('add_results', assessment_id=assessment.id)

        elif "add_results" in request.POST:
            if batch.status not in ["DRAFT", "FLAGGED"]:
                messages.error(request, "Results are locked after submission for verification.")
                return redirect('add_results', assessment_id=assessment.id)

            if batch.status == "FLAGGED":
                batch.rejection_reason = None
                batch.save(update_fields=["rejection_reason"])
                Result.objects.filter(batch=batch).update(status="DRAFT")

            with transaction.atomic():
                bulk_results = []
                for student in students_without_results:
                    score_raw = request.POST.get(f'score_{student.id}')
                    if score_raw:
                        try:
                            score = _parse_score_or_raise(score_raw, assessment, str(student))
                            bulk_results.append(Result(
                                assessment=assessment,
                                student=student,
                                score=score,
                                batch=batch,
                                status="DRAFT",
                            ))
                        except ValueError as exc:
                            messages.error(request, str(exc))
                            return redirect('add_results', assessment_id=assessment.id)

                try:
                    Result.objects.bulk_create(bulk_results)
                except IntegrityError:
                    messages.error(
                        request,
                        "Could not add results. Duplicate rows exist for one or more students in this assessment.",
                    )
                    return redirect('add_results', assessment_id=assessment.id)
                messages.success(request, "New results added successfully!")

                total_students = len(students)
                total_results = Result.objects.filter(
                    assessment=assessment,
                    student__is_active=True,
                ).count()
                if total_students > 0 and total_students == total_results:
                    messages.warning(
                        request,
                        "All results are now entered. Submit this batch for verification.",
                    )
            return redirect('add_results', assessment_id=assessment.id)

        elif "submit_batch" in request.POST:
            gate_errors = _submission_gate_errors(assessment, students)
            if gate_errors:
                for gate_error in gate_errors:
                    messages.error(request, gate_error)
                return redirect('add_results', assessment_id=assessment.id)

            logger.info(
                "submit_batch clicked: assessment_id=%s batch_id=%s status=%s results=%s",
                assessment.id,
                batch.id,
                batch.status,
                Result.objects.filter(assessment=assessment, student__is_active=True).count(),
            )
            submitted_batch, sample_count, ok = submit_batch_for_verification(assessment, request.user)
            logger.info(
                "submit_batch result: ok=%s batch_id=%s status=%s sample_count=%s",
                ok,
                submitted_batch.id if submitted_batch else None,
                getattr(submitted_batch, "status", None),
                sample_count,
            )
            if ok:
                messages.success(request, f"Batch submitted for verification. {sample_count} samples selected.")
            else:
                messages.error(request, "Batch already submitted or no results to submit.")
            return redirect('add_results', assessment_id=assessment.id)

        elif "unlock_batch" in request.POST:
            if batch.status == "VERIFIED":
                reset_batch_to_draft(batch, request.user)
                messages.success(request, "Batch unlocked and reset to Draft.")
            else:
                messages.error(request, "Only verified batches can be unlocked.")
            return redirect('add_results', assessment_id=assessment.id)

    # Render the results form
    effective_role = _effective_role_name(request)
    role_key = (effective_role or "").strip().lower()
    can_edit_results = (
        request.user.is_superuser
        or role_key in {"admin", "teacher"}
    )
    can_unlock_batch = (
        request.user.is_superuser
        or role_key in {"admin", "director of studies", "dos"}
    )
    can_view_verification_queue = (
        request.user.is_superuser
        or role_key in {"admin", "director of studies", "dos"}
    )

    if role_key in {"director of studies", "dos"}:
        return redirect('verification_queue', assessment_id=assessment.id)

    is_editable_batch = can_edit_results and batch.status in ["DRAFT", "FLAGGED"]

    results_count = existing_results.count()
    total_students = len(students)
    missing_count = max(total_students - results_count, 0)
    grading_bands = list(
        GradingSystem.objects.order_by("min_score").values("min_score", "max_score", "grade", "points")
    )
    for band in grading_bands:
        band["min_score"] = float(band["min_score"])
        band["max_score"] = float(band["max_score"])
        band["points"] = float(band["points"])

    existing_results_by_student = {item.student_id: item for item in existing_results}
    entry_rows = []
    for index, student in enumerate(students, start=1):
        result = existing_results_by_student.get(student.id)
        audit_label = "-"
        if result and result.batch and result.batch.submitted_by:
            audit_label = f"{result.batch.submitted_by.username}"
        elif result:
            audit_label = "Draft entry"

        entry_rows.append(
            {
                "index": index,
                "student": student,
                "result": result,
                "score": result.score if result else "",
                "grade": result.grade if result else "-",
                "points": result.points if result else "-",
                "status": result.status if result else "MISSING",
                "audit": audit_label,
            }
        )

    context = {
        'assessment': assessment,
        'students': students,
        'entry_rows': entry_rows,
        'students_without_results': students_without_results,
        'existing_results': existing_results,
        'current_mode': current_mode,
        'batch': batch,
        'all_results_entered': all_results_entered,
        'total_students': total_students,
        'results_count': results_count,
        'missing_results_count': missing_count,
        'teacher_label': request.user.get_full_name() or request.user.username,
        'entry_date': timezone.localdate(),
        'can_edit_results': can_edit_results,
        'is_editable_batch': is_editable_batch,
        'can_unlock_batch': can_unlock_batch,
        'can_view_verification_queue': can_view_verification_queue,
        'grading_bands': grading_bands,
    }
    return render(request, 'results/add_results_page.html', context)

@login_required
def edit_results_view(request, assessment_id=None, student_id=None):
    if not assessment_id:
        return redirect('class_assessment_list')

    assessment = get_object_or_404(Assessment, id=assessment_id)
    denial_reason = _mark_entry_denial_reason(request, assessment)
    if denial_reason:
        messages.error(request, denial_reason)
        return redirect('list_assessments', class_id=assessment.academic_class_id)
    academic_class = assessment.academic_class
    class_registers = ClassRegister.objects.filter(
        academic_class_stream__academic_class=academic_class,
        student__is_active=True,
    )
    all_students = [register.student for register in class_registers]

    if student_id:
        student = get_object_or_404(Student, id=student_id, is_active=True)
        results = Result.objects.filter(
            assessment=assessment,
            student=student,
            student__is_active=True,
        )
        form_students = [student]
    else:
        results = Result.objects.filter(assessment=assessment, student__is_active=True)
        form_students = all_students

    ResultFormSet = modelformset_factory(Result, form=ResultForm, extra=0)

    if request.method == "POST":
        formset = ResultFormSet(request.POST, queryset=results)
        if formset.is_valid():
            if assessment.result_batch and assessment.result_batch.status != "DRAFT":
                messages.error(request, "Results are locked after submission for verification.")
                return redirect('add_results', assessment_id=assessment_id)
            with transaction.atomic():
                for form in formset:
                    form.instance.assessment = assessment
                    if student_id:
                        form.instance.student = student
                formset.save()
                messages.success(request, "Results updated successfully!")
                return redirect('add_results', assessment_id=assessment_id)
        else:
            messages.error(request, "There was a problem with your input. Please check the form.")
    else:
        formset = ResultFormSet(queryset=results)

    zipped_forms = zip(formset.forms, form_students)
    current_mode = ResultModeSetting.get_mode()
    batch = ensure_batch_for_assessment(assessment)

    context = {
        'assessment': assessment,
        'formset': formset,
        'zipped_forms': zipped_forms,
        'current_mode': current_mode,
        'batch': batch,
    }
    return render(request, 'results/edit_results_page.html', context)



@login_required
def class_assessment_list_view(request):
    selected_year_id = request.GET.get('year_id')
    selected_term_id = request.GET.get('term_id')

    # Load filter options
    academic_years = AcademicYear.objects.all().order_by('-id')

    # Default selections: current academic year and current term (if not provided)
    if not selected_year_id:
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            selected_year_id = str(current_year.id)

    if selected_year_id and not selected_term_id:
        current_term = Term.objects.filter(academic_year_id=selected_year_id, is_current=True).first()
        if current_term:
            selected_term_id = str(current_term.id)

    # Terms list should always show for the currently selected year (or current year by default)
    if selected_year_id:
        terms = Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date')
    else:
        # Fallback: show all terms if no year context found
        terms = Term.objects.all().order_by('academic_year__id', 'start_date')

    # Role-aware class listing:
    # - Admins see all classes (unless they explicitly switch to Teacher/Class Teacher session role)
    # - Teachers see only classes where they are allocated (subject teacher) or class teacher in any term
    try:
        staff_account = StaffAccount.objects.get(user=request.user)
    except StaffAccount.DoesNotExist:
        staff_account = None

    session_role = request.session.get('active_role_name')
    role_name = session_role if session_role else (staff_account.role.name if staff_account and getattr(staff_account, "role", None) else None)
    role_key = (role_name or '').strip().lower()

    if request.user.is_superuser and role_key not in {'teacher', 'class teacher', 'class_teacher'}:
        academic_classes = AcademicClass.objects.all()
    elif role_key in {'teacher', 'class teacher', 'class_teacher'} and staff_account and getattr(staff_account, 'staff', None):
        # Subject-teacher allocations (any term) -> AcademicClass ids
        allocated_academic_class_ids = ClassSubjectAllocation.objects.filter(
            subject_teacher=staff_account.staff
        ).values_list('academic_class_stream__academic_class_id', flat=True).distinct()
        # Class-teacher assignments (any term) -> AcademicClass ids
        class_teacher_ac_ids = AcademicClassStream.objects.filter(
            class_teacher=staff_account.staff
        ).values_list('academic_class_id', flat=True).distinct()

        academic_classes = AcademicClass.objects.filter(
            Q(id__in=class_teacher_ac_ids) | Q(id__in=allocated_academic_class_ids)
        )
    else:
        academic_classes = AcademicClass.objects.none()

    # Apply current filters
    if selected_year_id:
        academic_classes = academic_classes.filter(academic_year_id=selected_year_id)
    if selected_term_id:
        academic_classes = academic_classes.filter(term_id=selected_term_id)
    academic_classes = academic_classes.distinct().order_by(
        'academic_year__id', 'term__start_date', 'Class__name'
    )

    # Restricted mode: no fallback to all classes for teachers

    is_teacher_role = role_key in {'teacher', 'class teacher', 'class_teacher'}

    context = {
        'academic_classes': academic_classes,
        'academic_years': academic_years,
        'terms': terms,
        'selected_year_id': str(selected_year_id or ''),
        'selected_term_id': str(selected_term_id or ''),
        'is_teacher_role': is_teacher_role,
    }
    return render(request, 'results/class_assessments.html', context)


#List of Assessments basing on specific academic_class
@login_required
def list_assessments_view(request, class_id):
  
    academic_class = get_object_or_404(AcademicClass, id=class_id)
    staff_account = StaffAccount.objects.filter(user=request.user).first()

    # Base queryset by authorization
    # If the current role is Teacher/Class Teacher, enforce restricted view even if the user is superuser
    if request.user.is_superuser and not StaffAccount.objects.filter(user=request.user, role__name__in=["Teacher","Class Teacher"]).exists():
        base_qs = Assessment.objects.filter(academic_class=academic_class)
        role_name = "Admin"
    else:
        session_role = request.session.get('active_role_name')
        role_name = session_role if session_role else (staff_account.role.name if staff_account and getattr(staff_account, "role", None) else None)
        role_key = (role_name or "").strip().lower()
        if staff_account or role_key:
            # Teachers: only their subject assessments; Class Teachers: all assessments in the class
            if role_key in {"class teacher", "class_teacher"}:
                is_class_teacher = AcademicClassStream.objects.filter(
                    class_teacher=staff_account.staff,
                    academic_class=academic_class,
                ).exists()
                if is_class_teacher:
                    base_qs = Assessment.objects.filter(academic_class=academic_class)
                else:
                    base_qs = Assessment.objects.none()
            elif role_key == "teacher":
                teacher_allocations = ClassSubjectAllocation.objects.filter(
                    subject_teacher=staff_account.staff,
                    academic_class_stream__academic_class=academic_class,
                )
                subject_ids = list(teacher_allocations.values_list('subject', flat=True))
                if subject_ids:
                    base_qs = Assessment.objects.filter(
                        academic_class=academic_class,
                        subject__in=subject_ids
                    )
                else:
                    base_qs = Assessment.objects.none()
            else:
                base_qs = Assessment.objects.none()
        else:
            base_qs = Assessment.objects.none()


    # --- Robust filters ---
    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    subject_id = to_int(request.GET.get('subject_id'))
    assessment_type_id = to_int(request.GET.get('assessment_type_id'))
    date_from_raw = request.GET.get('date_from')
    date_to_raw = request.GET.get('date_to')

    def parse_date(value):
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except (ValueError, TypeError):
                continue
        return None

    assessments = base_qs
    if subject_id:
        assessments = assessments.filter(subject_id=subject_id)
    if assessment_type_id:
        assessments = assessments.filter(assessment_type_id=assessment_type_id)

    df = parse_date(date_from_raw)
    dt = parse_date(date_to_raw)
    if df:
        assessments = assessments.filter(date__gte=df)
    if dt:
        assessments = assessments.filter(date__lte=dt)

    
    subjects = Subject.objects.filter(assessments__academic_class=academic_class).distinct().order_by('name')
    assessment_types = AssessmentType.objects.filter(assessments__academic_class=academic_class).distinct().order_by('name')

    
    # Normalize role name for case-insensitive checks
    role_key = (role_name or "").strip().lower()
    is_dos_user = False
    if role_key in {"director of studies", "dos"}:
        is_dos_user = True
    elif staff_account and getattr(staff_account, "role", None):
        if (staff_account.role.name or "").strip().lower() in {"director of studies", "dos"}:
            is_dos_user = True
    elif staff_account and getattr(staff_account, "staff", None):
        if staff_account.staff.roles.filter(name__in=["Director of Studies", "DOS"]).exists():
            is_dos_user = True

    allowed_roles = {
        "teacher",
        "head master",
        "headmaster",
        "head teacher",
        "headteacher",
        "director of studies",
        "dos",
        "admin",
    }
    status_filter = (request.GET.get("status") or "").strip().lower()
    if status_filter == "draft":
        assessments = assessments.filter(Q(result_batch__status="DRAFT") | Q(result_batch__isnull=True))
    is_teacher_role = role_key in {"teacher", "class teacher", "class_teacher"}
    if not is_teacher_role and staff_account and getattr(staff_account, "staff", None):
        if staff_account.staff.roles.filter(name__in=["Teacher", "Class Teacher"]).exists():
            is_teacher_role = True

    can_add_results = request.user.is_superuser or (role_key in allowed_roles)
    can_view_verification_queue = request.user.is_superuser or is_dos_user or role_key in {"admin"}

    assessments = assessments.select_related('result_batch')

    return render(request, 'results/list_assessments.html', {
        'assessments': assessments.order_by('-date', 'assessment_type__name', 'subject__name'),
        'academic_class': academic_class,
        'subjects': subjects,
        'assessment_types': assessment_types,
        'selected_subject_id': str(subject_id or ''),
        'selected_assessment_type_id': str(assessment_type_id or ''),
        'date_from': date_from_raw or '',
        'date_to': date_to_raw or '',
        'status_filter': status_filter,
        'can_add_results': can_add_results,
        'can_view_verification_queue': can_view_verification_queue,
        'role_key': role_key,
        'is_teacher_role': is_teacher_role,
    })



#Grading System
def _get_effective_role_name(request):
    return _effective_role_name(request)


def _can_manage_grading_systems(request):
    if request.user.is_superuser:
        return True
    return _effective_role_key(request) in VERIFICATION_ROLE_KEYS


@login_required
def grading_system_view(request):
    can_manage_grading = _can_manage_grading_systems(request)

    if request.method == "POST" and not can_manage_grading:
        messages.error(request, "Only Admin or Director of Studies can add grading systems.")
        return redirect("add_grading_system_page")

    if request.method == "POST":
        grading_form = GradingSystemForm(request.POST)
        if grading_form.is_valid():
            grading_form.save()
            messages.success(request, SUCCESS_ADD_MESSAGE)
        else:
            messages.error(request, FAILURE_MESSAGE)

    
    grading_form = GradingSystemForm()

    grading_systems = GradingSystem.objects.all()  

    context = {
        'grading_form': grading_form,
        'grading_systems': grading_systems,
        'is_dos': can_manage_grading,
        'can_manage_grading': can_manage_grading,
    }

    return render(request, 'results/grading_system.html', context)

#Edit grading system
@login_required
def edit_grading_system_view(request, id):
    if not _can_manage_grading_systems(request):
        messages.error(request, "Only Admin or Director of Studies can edit grading systems.")
        return redirect("add_grading_system_page")

    grading_system = get_model_record(GradingSystem,id)

    if request.method == "POST":
        grading_form = GradingSystemForm(request.POST, instance=grading_system)
        
        if grading_form.is_valid():
            grading_form.save()
            messages.success(request, SUCCESS_ADD_MESSAGE)
        else:
            messages.error(request, FAILURE_MESSAGE)
        return redirect("add_grading_system_page")

    else:
        grading_form = GradingSystemForm(instance=grading_system)
    
    context = {
        "grading_form": grading_form,
        "grading_system": grading_system
    }
    
    return render(request, 'results/edit_grading_system.html', context)


@login_required
def delete_grading_system_view(request, id):
    if not _can_manage_grading_systems(request):
        messages.error(request, "Only Admin or Director of Studies can delete grading systems.")
        return redirect("add_grading_system_page")

    grading_system = get_object_or_404(GradingSystem, pk=id)
    
    grading_system.delete()
    messages.success(request, DELETE_MESSAGE)
    return redirect("add_grading_system_page")

@login_required
def assessment_list_view(request):
    assessments = Assessment.objects.all()
    context = {
        'assessments': assessments,
    }
    return render(request, 'results/list_assessments.html', context)

@login_required
def exam_timetable_view(request):
    selected_year_id = request.GET.get("year_id")
    selected_term_id = request.GET.get("term_id")
    selected_class_id = request.GET.get("class_id")
    selected_assessment_type_id = request.GET.get("assessment_type_id")

    assessments = Assessment.objects.select_related(
        "academic_class",
        "assessment_type",
        "subject",
        "academic_class__academic_year",
        "academic_class__term",
    )

    if selected_year_id:
        assessments = assessments.filter(academic_class__academic_year_id=selected_year_id)
    if selected_term_id:
        assessments = assessments.filter(academic_class__term_id=selected_term_id)
    if selected_class_id:
        assessments = assessments.filter(academic_class__Class_id=selected_class_id)
    if selected_assessment_type_id:
        assessments = assessments.filter(assessment_type_id=selected_assessment_type_id)

    assessments = assessments.order_by("date", "academic_class__Class__name", "subject__name")

    assessments_by_date = defaultdict(list)
    for assessment in assessments:
        assessments_by_date[assessment.date].append(assessment)

    academic_years = AcademicYear.objects.all().order_by("-id")
    terms = Term.objects.all().order_by("term")
    classes = Class.objects.all().order_by("name")
    assessment_types = AssessmentType.objects.all().order_by("name")

    selected_year_label = AcademicYear.objects.filter(id=selected_year_id).first() if selected_year_id else None
    selected_term_label = Term.objects.filter(id=selected_term_id).first() if selected_term_id else None
    selected_class_label = Class.objects.filter(id=selected_class_id).first() if selected_class_id else None
    selected_assessment_type_label = AssessmentType.objects.filter(id=selected_assessment_type_id).first() if selected_assessment_type_id else None

    context = {
        "assessments_by_date": dict(assessments_by_date),
        "academic_years": academic_years,
        "terms": terms,
        "classes": classes,
        "assessment_types": assessment_types,
        "selected_year_id": str(selected_year_id or ""),
        "selected_term_id": str(selected_term_id or ""),
        "selected_class_id": str(selected_class_id or ""),
        "selected_assessment_type_id": str(selected_assessment_type_id or ""),
        "selected_year_label": selected_year_label.academic_year if selected_year_label else "All",
        "selected_term_label": selected_term_label.term if selected_term_label else "All",
        "selected_class_label": selected_class_label.name if selected_class_label else "All",
        "selected_assessment_type_label": selected_assessment_type_label.name if selected_assessment_type_label else "All",
        "generated_at": timezone.now(),
    }

    if request.GET.get("print") == "1":
        school = SchoolSetting.load()
        context.update({"school": school})
        html = render_to_string("results/exam_timetable_print.html", context)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = "inline; filename=exam_timetable.pdf"
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse("PDF generation error", status=500)
        return response

    return render(request, "results/exam_timetable.html", context)


@login_required
def add_assessment_view(request):
    from app.models.accounts import StaffAccount
    from app.models.classes import AcademicClassStream
    # ---- Helpers ----
    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    def parse_date(value):
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except (ValueError, TypeError):
                continue
        return None

    # ---- Read filters from GET ----
    selected_year_id = to_int(request.GET.get('year_id'))
    selected_term_id = to_int(request.GET.get('term_id'))
    selected_class_id = to_int(request.GET.get('class_id'))
    subject_id = to_int(request.GET.get('subject_id'))
    assessment_type_id = to_int(request.GET.get('assessment_type_id'))
    date_from_raw = request.GET.get('date_from')
    date_to_raw = request.GET.get('date_to')
    is_done_raw = request.GET.get('is_done')  # 'yes' | 'no' | None

    # ---- Role + class scope ----
    active_role = request.session.get("active_role_name")
    staff_account = StaffAccount.objects.filter(user=request.user).select_related("staff", "role").first()
    role_name = staff_account.role.name if staff_account and staff_account.role else None
    is_class_teacher = (active_role == "Class Teacher" or role_name == "Class Teacher")
    class_ids = []
    if is_class_teacher and staff_account and staff_account.staff:
        class_ids = list(
            AcademicClassStream.objects.filter(class_teacher=staff_account.staff)
            .values_list("academic_class_id", flat=True)
            .distinct()
        )

    # ---- Options for selects (Gentellella filter bar) ----
    academic_years = AcademicYear.objects.all().order_by('-id')
    if selected_year_id:
        terms = Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date')
    else:
        terms = Term.objects.all().order_by('academic_year__id', 'start_date')
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    assessment_types = AssessmentType.objects.all().order_by('name')

    # ---- Base queryset + filters ----
    assessments = (
        Assessment.objects
        .select_related('academic_class', 'assessment_type', 'subject')
        .all()
    )

    if is_class_teacher:
        if class_ids:
            assessments = assessments.filter(academic_class_id__in=class_ids)
        else:
            assessments = assessments.none()

    if selected_year_id:
        assessments = assessments.filter(academic_class__academic_year_id=selected_year_id)
    if selected_term_id:
        assessments = assessments.filter(academic_class__term_id=selected_term_id)
    if selected_class_id:
        assessments = assessments.filter(academic_class__Class_id=selected_class_id)
    if subject_id:
        assessments = assessments.filter(subject_id=subject_id)
    if assessment_type_id:
        assessments = assessments.filter(assessment_type_id=assessment_type_id)

    df = parse_date(date_from_raw)
    dt = parse_date(date_to_raw)
    if df:
        assessments = assessments.filter(date__gte=df)
    if dt:
        assessments = assessments.filter(date__lte=dt)

    if is_done_raw in ('yes', 'no'):
        assessments = assessments.filter(is_done=(is_done_raw == 'yes'))

    assessments = assessments.order_by('-date', 'assessment_type__name', 'subject__name', '-id')

    # ---- Verification Status Filter ----
    verification_status = request.GET.get('verification_status')
    if verification_status:
        if verification_status == "PENDING_OR_DRAFT":
            assessments = assessments.filter(result_batch__status__in=["PENDING", "DRAFT"])
        else:
            assessments = assessments.filter(result_batch__status=verification_status)

    # ---- Create assessment (modal submit) ----
    if request.method == "POST":
        if request.POST.get("bulk_subjects") == "1":
            bulk_form = BulkAssessmentForm(request.POST)
            form = AssessmentForm()
            if is_class_teacher:
                bulk_form.fields["academic_class"].queryset = AcademicClass.objects.filter(id__in=class_ids)
            if bulk_form.is_valid():
                academic_class = bulk_form.cleaned_data["academic_class"]
                if is_class_teacher and academic_class.id not in class_ids:
                    messages.error(request, "You can only create assessments for your assigned class.")
                    return redirect('assessment_create')
                assessment_type = bulk_form.cleaned_data["assessment_type"]
                subjects = bulk_form.cleaned_data["subjects"]
                date = bulk_form.cleaned_data["date"]
                out_of = bulk_form.cleaned_data["out_of"]
                is_done = bulk_form.cleaned_data["is_done"]

                created = 0
                skipped = 0
                for subject in subjects:
                    _, was_created = Assessment.objects.get_or_create(
                        academic_class=academic_class,
                        assessment_type=assessment_type,
                        subject=subject,
                        defaults={
                            "date": date,
                            "out_of": out_of,
                            "is_done": is_done,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        skipped += 1

                if created:
                    messages.success(request, f"Created {created} assessment(s).")
                if skipped:
                    messages.warning(request, f"Skipped {skipped} existing assessment(s).")
                return redirect('assessment_create')
            else:
                messages.error(request, FAILURE_MESSAGE)
        else:
            form = AssessmentForm(request.POST)
            bulk_form = BulkAssessmentForm()
            if is_class_teacher:
                form.fields["academic_class"].queryset = AcademicClass.objects.filter(id__in=class_ids)
            if form.is_valid():
                if is_class_teacher and form.cleaned_data["academic_class"].id not in class_ids:
                    messages.error(request, "You can only create assessments for your assigned class.")
                    return redirect('assessment_create')
                try:
                    form.save()
                    messages.success(request, SUCCESS_ADD_MESSAGE)
                    return redirect('assessment_create')
                except IntegrityError:
                    form.add_error(None, "An assessment for this Class, Type and Subject already exists.")
                    messages.error(request, "Duplicate assessment for this Class, Type and Subject.")
            else:
                messages.error(request, FAILURE_MESSAGE)
    else:
        form = AssessmentForm()
        bulk_form = BulkAssessmentForm()
        if is_class_teacher:
            form.fields["academic_class"].queryset = AcademicClass.objects.filter(id__in=class_ids)
            bulk_form.fields["academic_class"].queryset = AcademicClass.objects.filter(id__in=class_ids)

    context = {
        'form': form,
        'bulk_form': bulk_form,
        'assessments': assessments,
        # Filter options and selections
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'subjects': subjects,
        'assessment_types': assessment_types,
        'selected_year_id': str(selected_year_id or ''),
        'selected_term_id': str(selected_term_id or ''),
        'selected_class_id': str(selected_class_id or ''),
        'selected_subject_id': str(subject_id or ''),
        'selected_assessment_type_id': str(assessment_type_id or ''),
        'date_from': date_from_raw or '',
        'date_to': date_to_raw or '',
        'selected_is_done': is_done_raw or '',
        'verification_status': verification_status or '',
    }
    return render(request, 'results/add_assessment.html', context)

@login_required
def edit_assessment(request, id):
    assessment = get_model_record(Assessment,id)
    if request.method == "POST":
        form = AssessmentForm(request.POST, instance=assessment)
        if form.is_valid():
            form.save()
            messages.success(request,SUCCESS_ADD_MESSAGE)
            return redirect('assessment_create')
            
        else:
            messages.error(request,FAILURE_MESSAGE)
        
    else:
        form = AssessmentForm(instance=assessment)
    
    context = {
        'form': form,
        'assessment':assessment
    }
    return render(request, 'results/edit_assessment.html', context)

@login_required
def delete_assessment_view(request,id):
    if request.method != "POST":
        messages.error(request, "Delete requests must be submitted via POST.")
        return redirect(assessment_view)

    assessment = get_model_record(Assessment,id)
    assessment.delete()
    messages.success(request, DELETE_MESSAGE)
    return HttpResponseRedirect(reverse('assessment_create'))



@login_required
def assesment_type_view(request):
    if request.method == "POST":
        assesment_type_form = AssesmentTypeForm(request.POST)
        
        if assesment_type_form.is_valid():
            assesment_type_form.save()
            messages.success(request, SUCCESS_ADD_MESSAGE)
        else:
         
            messages.error(request,FAILURE_MESSAGE)
    
    assesment_type_form = AssesmentTypeForm()
    
    assesment_type = AssessmentType.objects.all()
    
    context = {
        "form": assesment_type_form,
        "assesment_type": assesment_type
    }
    
    return render(request, "results/assesment_type.html", context)

@login_required
def edit_assesment_type(request,id):
    assesment_type = get_model_record(AssessmentType,id)
    if request.method =="POST":
       assesment_type_form = AssesmentTypeForm(request.POST,instance=assesment_type)
       if assesment_type_form.is_valid():
           assesment_type_form.save().save()
           messages.success(request,SUCCESS_ADD_MESSAGE)
           return redirect(assesment_type_view)
       else:
           messages.error(request,FAILURE_MESSAGE)
           
    else:
        assesment_type_form = AssesmentTypeForm(instance=assesment_type)
    context={
        "form":assesment_type_form,
        "assesment_type":assesment_type
    } 
    return render(request,"results/edit_assesment_type.html",context)

@login_required
def delete_assesment_view(request, id):
    if request.method != "POST":
        messages.error(request, "Delete requests must be submitted via POST.")
        return redirect(assesment_type_view)

    assesment_type = AssessmentType.objects.get(pk=id)
    
    assesment_type.delete()
    messages.success(request, DELETE_MESSAGE)
    
    return redirect(assesment_type_view)

@login_required
def class_result_filter_view(request):
    years = AcademicYear.objects.all()
    academic_class_streams = AcademicClassStream.objects.none()

    # Ensure user has a staff account
    staff_account = getattr(request.user, "staff_account", None)

    if staff_account:
        role_name = staff_account.role.name
        active_role = request.session.get("active_role_name")

        if role_name in ["Admin", "Head master", "Director of Studies"] and active_role != "Class Teacher":
            # Full access
            academic_class_streams = AcademicClassStream.objects.select_related(
                'academic_class', 'stream', 'academic_class__academic_year'
            ).all()
        
        elif role_name == "Class Teacher" or active_role == "Class Teacher":
            # Restricted to streams assigned to them
            academic_class_streams = AcademicClassStream.objects.select_related(
                'academic_class', 'stream', 'academic_class__academic_year'
            ).filter(class_teacher=staff_account.staff)

        else:
            # e.g. Bursar, Support Staff → No class streams
            academic_class_streams = AcademicClassStream.objects.none()

    # Get selected parameters from GET request
    selected_year = request.GET.get('year_id')
    selected_class_stream = request.GET.get('class_stream_id')
    selected_term = request.GET.get('term_id')

    # Terms for the selected academic year
    terms = Term.objects.filter(academic_year_id=selected_year).order_by('start_date') if selected_year else Term.objects.none()

    # Apply filters and deduplicate class streams to prevent duplicates in the dropdown
    if academic_class_streams is not None:
        if selected_year:
            academic_class_streams = academic_class_streams.filter(
                academic_class__academic_year_id=selected_year
            )
        if selected_term:
            academic_class_streams = academic_class_streams.filter(
                academic_class__term_id=selected_term
            )
        academic_class_streams = academic_class_streams.order_by(
            'academic_class__Class__name',
            'stream__stream',
            'academic_class__academic_year__academic_year',
            'academic_class__id'
        ).distinct()

    students = Student.objects.none()
    no_students_message = None

    if selected_year and selected_class_stream:
        class_registers = ClassRegister.objects.filter(
            academic_class_stream_id=selected_class_stream,
            academic_class_stream__academic_class__academic_year_id=selected_year,
            student__is_active=True,
        )
        # If a term is selected, further scope by term
        if selected_term:
            class_registers = class_registers.filter(
                academic_class_stream__academic_class__term_id=selected_term
            )
        students = Student.objects.filter(
            is_active=True,
            id__in=class_registers.values('student_id'),
        ).order_by('student_name')

        if not students.exists():
            no_students_message = "No students found matching your criteria."

    # Derive bulk print params (year, term, class) from selected class stream
    bulk_year_id = None
    bulk_term_id = None
    bulk_class_id = None
    selected_assessment_type = request.GET.get('assessment_type_id')

    try:
        if selected_year and selected_class_stream:
            cs = AcademicClassStream.objects.select_related(
                'academic_class__Class',
                'academic_class__term',
                'academic_class__academic_year'
            ).filter(id=selected_class_stream).first()
            if cs and cs.academic_class:
                ac = cs.academic_class
                bulk_year_id = ac.academic_year_id
                # Prefer explicitly selected term; fallback to stream's academic class term
                bulk_term_id = selected_term or getattr(ac.term, 'id', None)
                bulk_class_id = getattr(ac.Class, 'id', None)
    except Exception:
        bulk_year_id = bulk_year_id or None
        bulk_term_id = bulk_term_id or None
        bulk_class_id = bulk_class_id or None

    context = {
        'years': years,
        'academic_class_streams': academic_class_streams,
        'selected_year': selected_year,
        'selected_class_stream': selected_class_stream,
        'students': students,
        'no_students_message': no_students_message,
        # Term filter options and selection
        'terms': terms,
        'selected_term': selected_term or '',
        # Needed for "Bulk Mini Reports (Assessment Type)" on Class Stream filter page
        'assessment_types': AssessmentType.objects.all(),
        'selected_assessment_type': selected_assessment_type or '',
        # Derived ids for bulk printing link
        'bulk_year_id': str(bulk_year_id) if bulk_year_id else '',
        'bulk_term_id': str(bulk_term_id) if bulk_term_id else '',
        'bulk_class_id': str(bulk_class_id) if bulk_class_id else '',
    }
    return render(request, 'results/class_stream_filter.html', context)




@login_required
def student_performance_view(request, student_id):
    student = get_object_or_404(Student, id=student_id, is_active=True)
    terms = Term.objects.filter(academic_year=student.academic_year)  
    assessment_types = AssessmentType.objects.all()  
    selected_term = request.GET.get("term_id")
    selected_assessment = request.GET.get("assessment_type")
    
    # Filter assessments by term and assessment type if selected
    assessments = Result.objects.filter(
        student=student,
        student__is_active=True,
    ).select_related(
        'assessment__subject',
        'assessment__assessment_type',
        'assessment__academic_class__term'
    ).order_by('assessment__date')
    
    if selected_term:
        assessments = assessments.filter(assessment__academic_class__term_id=selected_term)
    if selected_assessment:
        assessments = assessments.filter(assessment__assessment_type_id=selected_assessment)
    
    academic_class = AcademicClass.objects.filter(
        Class=student.current_class,
        academic_year=student.academic_year,
        term=student.term
    ).first()
    
    academic_class_stream = AcademicClassStream.objects.filter(
        academic_class=academic_class,
        stream=student.stream
    ).first() if academic_class else None
    
    performance_metrics = get_performance_metrics(assessments)
    subject_averages = calculate_weighted_subject_averages(assessments)
    
    subject_progress = {}
    for subject in set(r.assessment.subject.name for r in assessments):
        # Changed to use assessment_type.name instead of date
        subject_scores = [(r.assessment.assessment_type.name, r.score) for r in assessments if r.assessment.subject.name == subject]
        subject_scores.sort(key=lambda x: x[0])  # Sort by assessment_type name (alphabetically)
        progress = subject_scores[-1][1] - subject_scores[0][1] if len(subject_scores) > 1 else 0
        subject_progress[subject] = {
            'scores': subject_scores,
            'progress': progress,
            'trend': 'up' if progress > 0 else 'down' if progress < 0 else 'stable'
        }
    
    combined_subject_data = []
    for subject_avg in subject_averages:
        subject_name = subject_avg['subject']
        progress_data = subject_progress.get(subject_name, {'progress': 0, 'trend': 'stable', 'scores': []})
        combined_subject_data.append({
            **subject_avg,
            'progress': progress_data['progress'],
            'trend': progress_data['trend'],
            'scores': progress_data['scores'],
        })
    
    highest_subject = max(combined_subject_data, key=lambda s: s['average'], default=None)
    lowest_subject = min(combined_subject_data, key=lambda s: s['average'], default=None)
    
    performance_data = [
        {
            'score': float(performance_metrics['average']),
            'label': 'Overall Average Score',
            'icon': 'calculator',
            'type': 'avg',
            'subject': None
        },
        {
            'score': float(highest_subject['average']) if highest_subject else 0,
            'label': 'Best Performing Subject',
            'icon': 'trophy',
            'type': 'high',
            'subject': highest_subject['subject'] if highest_subject else 'N/A'
        },
        {
            'score': float(lowest_subject['average']) if lowest_subject else 0,
            'label': 'Lowest Performing Subject',
            'icon': 'exclamation-triangle',
            'type': 'low',
            'subject': lowest_subject['subject'] if lowest_subject else 'N/A'
        },
    ]
    
    assessment_data, assessment_dates = [], []
    grouped_assessments = {}
    for assessment in assessments:
        key = (assessment.assessment.date, assessment.assessment.assessment_type.name)
        grouped_assessments.setdefault(key, []).append(assessment)
    
    for (date, assessment_type), results in grouped_assessments.items():
        best_subject = max(results, key=lambda x: x.score)
        assessment_dates.append(date)
        assessment_data.append({
            'date': date,
            'type': assessment_type,
            'best_subject': {
                'name': best_subject.assessment.subject.name,
                'score': best_subject.score,
            },
            'subjects': {r.assessment.subject.name: r.score for r in results}
        })
    
   
    selected_term_name = Term.objects.filter(id=selected_term).first().term if selected_term else "All Terms"
    selected_assessment_name = AssessmentType.objects.filter(id=selected_assessment).first().name if selected_assessment else "All Assessment Types"
    
    context = {
        "student": {
            "obj": student,
            "details": {
                "full_name": student.student_name,
                "registration_number": student.reg_no,
                "class_info": f"{student.current_class.name} {student.stream.stream}",
                "academic_year": student.academic_year,
                "birthdate": student.birthdate,
                "gender": student.get_gender_display(),
                "nationality": student.get_nationality_display(),
                "religion": student.get_religion_display(),
                "guardian": student.guardian,
                "relationship": student.relationship,
                "contact": student.contact,
                "email": getattr(student, 'email', 'N/A'),
                "address": student.address,
                "photo_url": student.photo.url if student.photo else '/static/images/default-student.jpg'
            }
        },
        "terms": terms,
        "assessment_types": assessment_types,
        "assessments": performance_metrics['ordered_assessments'],
        "performance_data": performance_data,
        "subject_averages": combined_subject_data,
        "selected_term": selected_term,
        "selected_term_name": selected_term_name,
        "selected_assessment": selected_assessment,
        "selected_assessment_name": selected_assessment_name,
        "assessment_data": assessment_data,
        "subject_progress": subject_progress,
    }
    
    return render(request, "results/student_performance.html", context)


@login_required
def student_assessment_type_report(request, student_id, assessment_type_id):
    student = get_object_or_404(Student, id=student_id, is_active=True)

    # Determine selected term or fallback to student's current term
    selected_term_id = request.GET.get('term_id', str(student.term.id) if getattr(student, 'term', None) else None)
    terms = Term.objects.filter(academic_year=student.academic_year) if student.academic_year else Term.objects.none()

    academic_class = None
    if selected_term_id:
        academic_class = (
            AcademicClass.objects.filter(
                class_streams__classregister__student=student,
                term_id=selected_term_id
            )
            .select_related('academic_year', 'Class', 'term')
            .distinct()
            .first()
        )

    # Build per-student assessment-type report context (re-usable for bulk)
    context = build_student_assessment_type_context(
        student,
        assessment_type_id,
        selected_term_id,
        academic_class=academic_class
    )

    # Order assessment types like term report (for consistent UI)
    assessment_order = Case(
        When(name__iexact="BEGINNING OF TERM", then=Value(1)),
        When(name__iexact="MID OF TERM", then=Value(2)),
        When(name__iexact="END OF TERM INTERNAL", then=Value(3)),
        When(name__iexact="END OF TERM EXTERNAL", then=Value(4)),
        default=Value(5),
        output_field=IntegerField()
    )
    assessment_types = AssessmentType.objects.all().order_by(assessment_order, 'name')

    # Augment context for template dropdowns (preserve any fallback term chosen in builder)
    context.update({
        'terms': terms,
        'assessment_types': assessment_types,
        'selected_term_id': context.get('selected_term_id'),
    })

    return render(request, 'results/student_assessment_report.html', context)


def build_student_assessment_type_context(student, assessment_type_id, selected_term_id, academic_class=None):
    """
    Builds the per-student report context for a specific assessment type.
    Preserves data and report format used by templates/results/student_assessment_report.html
    """
    school = SchoolSetting.load()
    assessment_type = get_object_or_404(AssessmentType, id=assessment_type_id)

    resolved_academic_class = academic_class
    if not resolved_academic_class and selected_term_id:
        resolved_academic_class = (
            AcademicClass.objects.filter(
                class_streams__classregister__student=student,
                term_id=selected_term_id
            )
            .select_related('academic_year', 'Class', 'term')
            .distinct()
            .first()
        )

    # Base queryset
    results = (
        Result.objects
        .filter(
            student=student,
            student__is_active=True,
            assessment__assessment_type=assessment_type,
        )
        .select_related(
            'assessment__subject',
            'assessment__assessment_type',
            'assessment__academic_class__term'
        )
    )
    if selected_term_id:
        results = results.filter(assessment__academic_class__term_id=selected_term_id)
    if resolved_academic_class:
        results = results.filter(assessment__academic_class=resolved_academic_class)

    # Handle fallback to latest term with results if none in selected term
    no_results_message = None
    term_label = "-"
    if not results.exists():
        term_ids_with_results = (
            Result.objects
            .filter(
                student=student,
                student__is_active=True,
                assessment__assessment_type=assessment_type,
            )
            .values_list('assessment__academic_class__term_id', flat=True)
            .distinct()
        )
        fallback_term = (
            Term.objects.filter(id__in=term_ids_with_results)
            .order_by('start_date')
            .last()
        )
        if fallback_term:
            selected_term_id = str(fallback_term.id)
            if not resolved_academic_class:
                resolved_academic_class = (
                    AcademicClass.objects.filter(
                        class_streams__classregister__student=student,
                        term_id=fallback_term.id
                    )
                    .select_related('academic_year', 'Class', 'term')
                    .distinct()
                    .first()
                )
            results = (
                Result.objects
                .filter(
                    student=student,
                    student__is_active=True,
                    assessment__assessment_type=assessment_type,
                    assessment__academic_class__term_id=fallback_term.id
                )
                .select_related(
                    'assessment__subject',
                    'assessment__assessment_type',
                    'assessment__academic_class__term'
                )
            )
            if resolved_academic_class:
                results = results.filter(assessment__academic_class=resolved_academic_class)
            term_label = fallback_term.term
            no_results_message = "No results in selected term. Showing latest term with data."
        else:
            no_results_message = "No results available for this assessment type."
    else:
        if selected_term_id:
            t = Term.objects.filter(id=selected_term_id).first()
            term_label = t.term if t else "-"

    # Group by subject and compute per-subject averages
    from decimal import Decimal
    subject_scores = {}
    for result in results:
        subject = result.assessment.subject.name
        if subject not in subject_scores:
            subject_scores[subject] = {'scores': [], 'total': Decimal('0.0'), 'count': 0}
        subject_scores[subject]['scores'].append(result)
        subject_scores[subject]['total'] += Decimal(str(result.score))
        subject_scores[subject]['count'] += 1

    summary = []
    for subject, data in subject_scores.items():
        avg = (data['total'] / data['count']).quantize(Decimal('0.01')) if data['count'] else Decimal('0.00')
        grade, points = get_grade_and_points(avg)
        summary.append({
            'subject': subject,
            'average': float(avg),
            'grade': grade,
            'points': points,
            'details': data['scores'],
        })

    # Totals and overall
    total_marks = sum(Decimal(str(r.score)) for r in results) if results else Decimal('0.00')
    total_aggregates = sum(item['points'] for item in summary) if summary else 0
    overall_average = (total_marks / len(results)).quantize(Decimal('0.01')) if results else Decimal('0.00')
    overall_grade, overall_points = get_grade_and_points(overall_average)
    subject_grades = {
        item['subject']: {
            'grade': item['grade'],
            'points': item['points']
        }
        for item in summary
    }
    if total_aggregates:
        selected_division, division_override_note = get_division_with_override(
            int(total_aggregates),
            subject_grades
        )
    else:
        selected_division, division_override_note = "-", None

    academic_year_source = (
        resolved_academic_class.academic_year
        if resolved_academic_class
        else student.academic_year
    )
    academic_year = academic_year_source.academic_year if academic_year_source else "-"

    # Signatures
    head_teacher_signature = Signature.objects.filter(position="HEAD TEACHER").first()
    class_teacher_signature = None
    try:
        academic_class = resolved_academic_class
        if not academic_class and selected_term_id:
            academic_class = AcademicClass.objects.filter(
                Class=student.current_class,
                academic_year=student.academic_year,
                term_id=selected_term_id
            ).first()

        if academic_class:
            class_stream = AcademicClassStream.objects.filter(
                academic_class=academic_class,
                stream=student.stream
            ).first()
            class_teacher_signature = class_stream.class_teacher_signature if class_stream else None
    except Exception:
        class_teacher_signature = None

    # Avoid duplicating the same image for both signatures
    try:
        if (
            class_teacher_signature
            and head_teacher_signature
            and getattr(head_teacher_signature, 'signature', None)
            and getattr(class_teacher_signature, 'name', None)
            and class_teacher_signature.name == head_teacher_signature.signature.name
        ):
            class_teacher_signature = None
    except Exception:
        pass

    return {
        'school': school,
        'student': student,
        'assessment_type': assessment_type,
        'summary': summary,
        'term': term_label,
        'academic_year': academic_year,
        'total_marks': float(total_marks),
        'total_aggregates': total_aggregates,
        'overall_average': float(overall_average),
        'overall_grade': overall_grade,
        'overall_points': overall_points,
        'selected_term_id': str(selected_term_id) if selected_term_id else None,
        'head_teacher_signature': head_teacher_signature,
        'class_teacher_signature': class_teacher_signature,
        'no_results_message': no_results_message,
        'selected_division': selected_division,
        'division_override_note': division_override_note,
    }

    
@login_required
def student_term_report(request, student_id):
    student = get_object_or_404(Student, id=student_id, is_active=True)
    report_format = request.GET.get('report_format') or 'standard'
    report_template_map = {
        'standard': 'results/student_term_report.html',
        'tahfiz': 'results/tahfiz/tahfiz_report.html',
    }
    if report_format not in report_template_map:
        report_format = 'standard'
    report_template = report_template_map[report_format]

    # Get all terms in the student's academic year
    terms = Term.objects.filter(
        academic_year=student.academic_year
    ).order_by('id')

    # Get selected term_id from query param or default
    selected_term_id = request.GET.get('term_id')
    if not selected_term_id:
        selected_term_id = (
            student.term.id if student.term
            else (terms.first().id if terms.exists() else None)
        )

    if not selected_term_id:
        messages.error(request, "No terms available for this student.")
        return redirect('student_performance_view', student_id=student_id)

    academic_class = None
    if selected_term_id:
        academic_class = (
            AcademicClass.objects.filter(
                class_streams__classregister__student=student,
                term_id=selected_term_id
            )
            .select_related('academic_year', 'Class', 'term')
            .distinct()
            .first()
        )

    context = build_student_report_context(
        student,
        selected_term_id,
        academic_class=academic_class
    )

    
    context['terms'] = terms
    context['selected_term_id'] = str(selected_term_id)
    context['selected_report_format'] = report_format

    return render(request, report_template, context)

def build_student_report_context(student, term_id, academic_class=None):
    school = SchoolSetting.load()
    term = get_object_or_404(Term, id=term_id)

    resolved_academic_class = academic_class
    if not resolved_academic_class:
        resolved_academic_class = (
            AcademicClass.objects.filter(
                class_streams__classregister__student=student,
                term_id=term_id
            )
            .select_related('academic_year', 'Class', 'term')
            .distinct()
            .first()
        )
    academic_year = (
        resolved_academic_class.academic_year
        if resolved_academic_class
        else term.academic_year
    )
    class_size = None
    if resolved_academic_class:
        class_size = (
            ClassRegister.objects
            .filter(
                academic_class_stream__academic_class=resolved_academic_class,
                student__is_active=True,
            )
            .values('student_id')
            .distinct()
            .count()
        )

    # ---- Custom order definition for assessment types ----
    assessment_order = Case(
        When(name__iexact="BEGINNING OF TERM", then=Value(1)),
        When(name__iexact="MID OF TERM", then=Value(2)),
        When(name__iexact="END OF TERM INTERNAL", then=Value(3)),
        When(name__iexact="END OF TERM EXTERNAL", then=Value(4)),
        default=Value(5),
        output_field=IntegerField()
    )

    results_order = Case(
        When(assessment__assessment_type__name__iexact="BEGINNING OF TERM", then=Value(1)),
        When(assessment__assessment_type__name__iexact="MID OF TERM", then=Value(2)),
        When(assessment__assessment_type__name__iexact="END OF TERM INTERNAL", then=Value(3)),
        When(assessment__assessment_type__name__iexact="END OF TERM EXTERNAL", then=Value(4)),
        default=Value(5),
        output_field=IntegerField()
    )

    # Fetch assessment types in the desired order
    assessment_types = AssessmentType.objects.all().order_by(assessment_order, 'name')

    # Fetch results ordered by subject and custom assessment order
    results = Result.objects.filter(
        student=student,
        student__is_active=True,
        assessment__academic_class__term_id=term_id,
    )
    if resolved_academic_class:
        results = results.filter(assessment__academic_class=resolved_academic_class)
    results = results.select_related(
        'assessment__subject',
        'assessment__assessment_type',
        'assessment__academic_class__term'
    ).order_by(
        'assessment__subject__name',
        results_order
    )

    # Prepare subject summaries and totals
    subject_summary = {}
    total_marks = Decimal('0.0')
    total_weight = Decimal('0.0')
    assessment_totals = {
        at.name: {'marks': Decimal('0.0'), 'points': Decimal('0.0'), 'count': 0}
        for at in assessment_types
    }

    for result in results:
        subject_name = result.assessment.subject.name or ''
        subject_desc = result.assessment.subject.description or ''
        subject_clean = (subject_name + " " + subject_desc).upper().replace(" ", "")
        exclude_from_totals = (
            'RELIGIOUSEDUCATION' in subject_clean or
            'READING' in subject_clean
        )

        assessment_type = result.assessment.assessment_type.name
        weight = Decimal(str(result.assessment.assessment_type.weight or 1))
        score = Decimal(str(result.score))

        if subject_name not in subject_summary:
            subject_summary[subject_name] = {
                'assessments': {},
                'total_score': Decimal('0.0'),
                'total_weight': Decimal('0.0')
            }

        grade, points = get_grade_and_points(score)
        subject_summary[subject_name]['assessments'][assessment_type] = {
            'score': float(score),
            'grade': grade,
            'points': points
        }
        subject_summary[subject_name]['total_score'] += score * weight
        subject_summary[subject_name]['total_weight'] += weight

        if not exclude_from_totals:
            total_marks += score * weight
            total_weight += weight

            if assessment_type in assessment_totals:
                assessment_totals[assessment_type]['marks'] += score
                assessment_totals[assessment_type]['points'] += Decimal(str(points))
                assessment_totals[assessment_type]['count'] += 1

    # Build report data per subject
    report_data = []
    for subject, data in subject_summary.items():
        avg = (
            data['total_score'] / data['total_weight']
        ).quantize(Decimal('0.01')) if data['total_weight'] else Decimal('0.00')
        grade, points = get_grade_and_points(avg)
        report_data.append({
            'subject': subject,
            'average': float(avg),
            'grade': grade,
            'points': points,
            'assessments': {
                at.name: data['assessments'].get(
                    at.name,
                    {'score': '-', 'grade': '-', 'points': '-'}
                )
                for at in assessment_types
            }
        })

    # Overall stats
    overall_average = (
        total_marks / total_weight
    ).quantize(Decimal('0.01')) if total_weight else Decimal('0.00')
    overall_grade, overall_points = get_grade_and_points(overall_average)

    subject_grades = {
        item['subject']: {
            'grade': item['grade'],
            'points': item['points']
        }
        for item in report_data
    }

    # Calculate assessment divisions (per assessment type)
    assessment_divisions = {}
    for at in assessment_types:
        total_points = int(assessment_totals[at.name]['points'])
        count = assessment_totals[at.name]['count']
        assessment_divisions[at.name] = (
            get_division(total_points, subject_grades) if count > 0 else "-"
        )

    total_aggregates = sum(
        assessment_totals[at.name]['points'] for at in assessment_types
        if assessment_totals[at.name]['count'] > 0
    )

    if total_aggregates:
        selected_division, division_override_note = get_division_with_override(
            int(total_aggregates),
            subject_grades
        )
    else:
        selected_division, division_override_note = "-", None

    academic_year_str = academic_year.academic_year if academic_year else "-"
    term_name = term.term if term else "-"
    class_display = "-"
    if resolved_academic_class and getattr(resolved_academic_class, "Class", None):
        class_display = (
            getattr(resolved_academic_class.Class, "code", None)
            or getattr(resolved_academic_class.Class, "name", None)
            or str(resolved_academic_class.Class)
        )
    elif getattr(student, "current_class", None):
        class_display = (
            getattr(student.current_class, "code", None)
            or getattr(student.current_class, "name", None)
            or str(student.current_class)
        )

    overall_average_float = float(overall_average) if overall_average is not None else 0.0
    if overall_average_float >= 85:
        discipline_rating = "Excellent"
        discipline_rating_ar = "ممتاز"
    elif overall_average_float >= 75:
        discipline_rating = "Very Good"
        discipline_rating_ar = "جيد جدا"
    elif overall_average_float >= 65:
        discipline_rating = "Good"
        discipline_rating_ar = "جيد"
    elif overall_average_float >= 50:
        discipline_rating = "Fair"
        discipline_rating_ar = "مقبول"
    else:
        discipline_rating = "Needs Improvement"
        discipline_rating_ar = "يحتاج إلى تحسين"

    # --- Find next term ---
    next_term = None
    if term and student.academic_year:
        next_term = Term.objects.filter(
            academic_year=student.academic_year,
            start_date__gt=term.start_date
        ).order_by('start_date').first()

    next_term_start_date = next_term.start_date if next_term else None
    next_term_name = next_term.term if next_term else None

    colspan = 2 + len(assessment_types) + 1

    # ---- Signatures ----
    head_teacher_signature = Signature.objects.filter(position="HEAD TEACHER").first()
    class_teacher_signature = None
    try:
        academic_class_obj = resolved_academic_class
        if not academic_class_obj:
            # Resolve the academic class for this student in the given term
            academic_class_obj = AcademicClass.objects.filter(
                Class=student.current_class,
                academic_year=student.academic_year,
                term=term
            ).first()
        if academic_class_obj:
            class_stream = AcademicClassStream.objects.filter(
                academic_class=academic_class_obj,
                stream=student.stream
            ).first()
            class_teacher_signature = class_stream.class_teacher_signature if class_stream else None
    except Exception:
        class_teacher_signature = None

    # Persisted report remarks
    remark = ReportRemark.objects.filter(student=student, term=term).first()
    class_teacher_remark = (remark.class_teacher_remark or "").strip() if remark else ""
    head_teacher_remark = (remark.head_teacher_remark or "").strip() if remark else ""

    # Print metadata
    printed_at = timezone.localtime()
    year_ref = academic_year.id if getattr(academic_year, "id", None) else "NA"
    term_ref = term.id if getattr(term, "id", None) else "NA"
    report_reference = f"TR-{student.id}-{year_ref}-{term_ref}"

    return {
        'school': school,
        'student': student,
        'report_data': report_data,
        'assessment_types': assessment_types,
        'term': term_name,
        'academic_year': academic_year_str,
        'class_display': class_display,
        'total_marks': float(total_marks),
        'overall_average': float(overall_average),
        'overall_grade': overall_grade,
        'overall_points': overall_points,
        'colspan': colspan,
        'assessment_totals': assessment_totals,
        'total_aggregates': int(total_aggregates),
        'assessment_divisions': assessment_divisions,
        'selected_division': selected_division,
        'division_override_note': division_override_note,
        'next_term_start_date': next_term_start_date,
        'next_term_name': next_term_name,
        'head_teacher_signature': head_teacher_signature,
        'class_teacher_signature': class_teacher_signature,
        'class_teacher_remark': class_teacher_remark,
        'head_teacher_remark': head_teacher_remark,
        'printed_at': printed_at,
        'report_reference': report_reference,
        'class_size': class_size,
        'discipline_rating': discipline_rating,
        'discipline_rating_ar': discipline_rating_ar,
    }


@login_required
def class_bulk_reports(request):
    academic_year_id = request.GET.get('academic_year_id')
    term_id = request.GET.get('term_id')
    class_id = request.GET.get('class_id')
    report_format = request.GET.get('report_format') or 'standard'

    report_template_map = {
        'standard': 'results/student_term_report.html',
        'tahfiz': 'results/tahfiz/tahfiz_report.html',
    }
    if report_format not in report_template_map:
        report_format = 'standard'
    report_template = report_template_map[report_format]

    if not (academic_year_id and term_id and class_id):
        messages.error(request, "Please select Academic Year, Term, and Class first.")
        return redirect('class_performance_summary')

    academic_class = get_object_or_404(
        AcademicClass,
        academic_year_id=academic_year_id,
        term_id=term_id,
        Class_id=class_id
    )

    # Use ClassRegister to get students who were enrolled in this academic class
    # This ensures promoted students still appear in historical reports
    student_ids = ClassRegister.objects.filter(
        academic_class_stream__academic_class=academic_class,
        student__is_active=True,
    ).values_list('student_id', flat=True).distinct()
    
    students = Student.objects.filter(is_active=True, id__in=student_ids).order_by('student_name')
    school = SchoolSetting.load()
   

    reports = [build_student_report_context(student, term_id, academic_class=academic_class) for student in students]
    head_teacher_signature = Signature.objects.filter(position="HEAD TEACHER").first()

    context = {
        'school': school,
        'reports': reports,
        'class_obj': academic_class,
        'head_teacher_signature': head_teacher_signature,
        'report_format': report_format,
        'report_template': report_template,
    }
    return render(request, 'results/class_bulk_reports.html', context)


@login_required
def class_assessment_type_bulk_reports(request):
   
    academic_year_id = request.GET.get('academic_year_id')
    term_id = request.GET.get('term_id')
    class_id = request.GET.get('class_id')
    assessment_type_id = request.GET.get('assessment_type_id')

    if not (academic_year_id and term_id and class_id and assessment_type_id):
        messages.error(request, "Please select Academic Year, Term, Class and Assessment Type first.")
        return redirect('class_performance_summary')

    academic_class = get_object_or_404(
        AcademicClass,
        academic_year_id=academic_year_id,
        term_id=term_id,
        Class_id=class_id
    )

    # Use ClassRegister to get students who were enrolled in this academic class
    # This ensures promoted students still appear in historical reports
    student_ids = ClassRegister.objects.filter(
        academic_class_stream__academic_class=academic_class,
        student__is_active=True,
    ).values_list('student_id', flat=True).distinct()
    
    students = Student.objects.filter(is_active=True, id__in=student_ids).order_by('student_name')

    school = SchoolSetting.load()
    assessment_type = get_object_or_404(AssessmentType, id=assessment_type_id)

    # Build per-student contexts (preserving template format)
    reports = [
        build_student_assessment_type_context(student, assessment_type_id, term_id, academic_class=academic_class)
        for student in students
    ]

    head_teacher_signature = Signature.objects.filter(position="HEAD TEACHER").first()

    context = {
        'school': school,
        'assessment_type': assessment_type,
        'reports': reports,
        'class_obj': academic_class,
        'head_teacher_signature': head_teacher_signature,
    }
    return render(request, 'results/class_assessment_type_bulk_reports.html', context)

@login_required
def class_performance_summary(request):
    academic_year_id = request.GET.get('academic_year_id')
    term_id = request.GET.get('term_id')
    class_id = request.GET.get('class_id')
    assessment_type_id = request.GET.get('assessment_type_id')
    report_format = request.GET.get('report_format') or 'standard'

    def to_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    academic_year_id = to_int(academic_year_id)
    term_id = to_int(term_id)
    class_id = to_int(class_id)
    assessment_type_id = to_int(assessment_type_id)

    if report_format not in {'standard', 'tahfiz'}:
        report_format = 'standard'

    academic_years = AcademicYear.objects.all()
    terms = Term.objects.all()
    classes = Class.objects.all()
    assessment_types = AssessmentType.objects.all()

    best_students = []
    subject_averages = []
    students_data = []
    subjects = []

    academic_class = None
    academic_class_exists = False

    if academic_year_id and term_id and class_id:
        try:
            academic_class = AcademicClass.objects.get(
                Class_id=class_id,
                academic_year_id=academic_year_id,
                term_id=term_id
            )
            academic_class_exists = True
        except AcademicClass.DoesNotExist:
            
            messages.warning(request, "No records found for the selected Academic Year, Term, and Class.")
            academic_class = None
            academic_class_exists = False

        if academic_class:
            # Optimize queries with select_related and prefetch_related
            results_qs = Result.objects.filter(assessment__academic_class=academic_class).select_related(
                'student', 'assessment__subject', 'assessment__assessment_type'
            )
            results_qs = results_qs.filter(student__is_active=True)
            if assessment_type_id:
                results_qs = results_qs.filter(assessment__assessment_type_id=assessment_type_id)

            # Get all students for this class with optimized query
            students = Student.objects.filter(
                current_class_id=class_id,
                is_active=True,
            ).select_related(
                'current_class', 'stream'
            ).prefetch_related('results')

            # Pre-calculate all metrics in fewer queries
            # Best students - optimized
            best_students = (
                results_qs
                .values('student__student_name', 'student__current_class__name')
                .annotate(average=Avg('score'))
                .order_by('-average')[:5]
            )

            # Subject averages - optimized
            subject_averages = (
                results_qs
                .values('assessment__subject__name')
                .annotate(avg_score=Avg('score'), best_score=Max('score'))
                .order_by('-avg_score')
            )

            # Build students data more efficiently
            students_dict = {student.id: student for student in students}
            subjects = Subject.objects.all().order_by('id')

            # Get all results for students in this class
            student_results = results_qs.values(
                'student_id', 'score', 'assessment__subject_id'
            )

            # Group results by student
            results_by_student = defaultdict(list)
            for result in student_results:
                results_by_student[result['student_id']].append(result)

            # Build students_data efficiently
            for student_id, student in students_dict.items():
                student_results_list = results_by_student.get(student_id, [])
                results = {}
                total_marks = 0
                total_agg = 0

                for subject in subjects:
                    res = next((r for r in student_results_list if r['assessment__subject_id'] == subject.id), None)
                    if res:
                        # Calculate points using the grading system
                        points = 0
                        try:
                            grading = GradingSystem.objects.filter(
                                min_score__lte=res['score'],
                                max_score__gte=res['score']
                            ).first()
                            points = grading.points if grading else 0
                        except:
                            points = 0

                        results[subject.id] = {
                            'marks': res['score'],
                            'agg': points,
                        }
                        total_marks += res['score']
                        total_agg += points

                students_data.append({
                    'student': student,
                    'results': results,
                    'total_marks': total_marks,
                    'total_agg': total_agg,
                    'division': get_division(total_agg) if total_agg else '-'
                })

    # Calculate additional KPIs and data for the dashboard
    total_students = 0
    average_score = 0
    pass_rate = 0
    top_performer = None
    current_term = None
    class_performance = []
    subject_performance = []
    assessment_type_performance = []
    top_students = []
    grade_distribution = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    gender_comparison = {'Male': 0, 'Female': 0}
    stream_comparison = {}
    subject_heatmap = {}
    performance_trends = []

    if academic_class:
        total_students = students.count()
        average_score = results_qs.aggregate(Avg('score'))['score__avg'] or 0
        pass_rate = (results_qs.filter(score__gte=70).count() / results_qs.count() * 100) if results_qs.count() > 0 else 0
        top_performer = results_qs.values('student__student_name').annotate(average=Avg('score')).order_by('-average').first()
        current_term = academic_class.term.term

        # Class Performance Overview (for this class)
        class_performance = [{
            'class': str(academic_class.Class.name),
            'average': float(average_score or 0),
            'students': int(total_students or 0)
        }]

        # Subject Performance Overview
        subject_performance = [
            {
                'subject': avg['assessment__subject__name'],
                'average': float(avg['avg_score'] or 0),
                'best_score': float(avg['best_score'] or 0),
            } for avg in subject_averages
        ]

        # Assessment Type Comparison
        assessment_type_performance = list(
            results_qs
            .values('assessment__assessment_type__name')
            .annotate(avg_score=Avg('score'))
            .order_by('-avg_score')
        )
        assessment_type_performance = [
            {
                'assessment__assessment_type__name': a['assessment__assessment_type__name'],
                'avg_score': float(a['avg_score'] or 0),
            }
            for a in assessment_type_performance
        ]

        # Top 10 Students
        top_students = (
            results_qs
            .values('student__student_name', 'student__current_class__name')
            .annotate(average=Avg('score'))
            .order_by('-average')[:10]
        )


        # Gender Performance Comparison - calculate average scores by gender
        gender_performance = results_qs.values('student__gender').annotate(
            avg_score=Avg('score'),
            student_count=Count('id')
        ).order_by('-avg_score')

        gender_comparison = {}
        for perf in gender_performance:
            gender = perf['student__gender']
            if gender == 'M':
                gender_comparison['Male'] = {
                    'average': float(perf['avg_score'] or 0),
                    'count': int(perf['student_count'] or 0)
                }
            elif gender == 'F':
                gender_comparison['Female'] = {
                    'average': float(perf['avg_score'] or 0),
                    'count': int(perf['student_count'] or 0)
                }

        # Stream comparison - optimized
        stream_stats = students.values('stream__stream').annotate(count=Count('id')).order_by('-count')
        stream_comparison = {stat['stream__stream'] or 'No Stream': stat['count'] for stat in stream_stats}

        # Subject Difficulty Heatmap - optimized with single query
        heatmap_data = results_qs.values('assessment__subject__name').annotate(avg_score=Avg('score')).order_by('assessment__subject__name')
        subject_heatmap = {item['assessment__subject__name']: item['avg_score'] for item in heatmap_data}

        # Performance Trends Over Terms (simplified, for this term only)
        performance_trends = [{
            'term': current_term,
            'average': float(average_score or 0)
        }]

    # Get academic year and term names for display
    selected_academic_year_name = ''
    selected_term_name = ''
    if academic_year_id:
        try:
            selected_academic_year_name = AcademicYear.objects.get(id=academic_year_id).academic_year
        except AcademicYear.DoesNotExist:
            selected_academic_year_name = ''
    if term_id:
        try:
            selected_term_name = Term.objects.get(id=term_id).term
        except Term.DoesNotExist:
            selected_term_name = ''

    context = {
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'assessment_types': assessment_types,
        'best_students': best_students,
        'subject_averages': subject_averages,
        'students_data': students_data,
        'subjects': subjects,
        'selected_academic_year': str(academic_year_id) if academic_year_id else '',
        'selected_term': str(term_id) if term_id else '',
        'selected_class': str(class_id) if class_id else '',
        'selected_assessment_type': str(assessment_type_id) if assessment_type_id else '',
        'selected_report_format': report_format,
        'selected_academic_year_name': selected_academic_year_name,
        'selected_term_name': selected_term_name,
        'academic_class_exists': academic_class_exists,
        # New dashboard data
        'total_students': total_students,
        'average_score': average_score,
        'pass_rate': pass_rate,
        'top_performer': top_performer,
        'current_term': current_term,
        'class_performance': class_performance,
        'subject_performance': subject_performance,
        'assessment_type_performance': assessment_type_performance,
        'top_students': top_students,
        'gender_comparison': gender_comparison,
        'stream_comparison': stream_comparison,
        'subject_heatmap': subject_heatmap,
        'performance_trends': performance_trends,
    }

    return render(request, 'results/class_performance_summary.html', context)



@login_required
def assessment_sheet_view(request):
    def q2(val, places="0.01"):
        """Quantize the value to specified decimal places."""
        return val.quantize(Decimal(places), rounding=ROUND_HALF_UP)

    def is_excluded_subject(name: str, desc: str = "") -> bool:
        text = f"{(name or '')} {(desc or '')}".upper().replace(" ", "")
        return "READING" in text or "RELIGIOUSEDUCATION" in text

    def get_division(aggregates):
        """Determine division based on total aggregates"""
        if aggregates <= 12:
            return 1
        elif aggregates <= 23:
            return 2
        elif aggregates <= 29:
            return 3
        elif aggregates <= 34:
            return 4
        else:
            return "U"

    def norm_key(s: str) -> str:
        """Normalize subject-like labels for reliable dict keys."""
        return (s or "").strip().upper()

    # Get all unique grades (classes)
    unique_grades = Class.objects.values_list("name", flat=True).distinct()
    selected_grade = request.GET.get("grade")

    # Get all academic years for the filter
    academic_years = AcademicYear.objects.all().order_by('-id')
    selected_year_id = request.GET.get("year_id")
    
    # Default to current academic year if not selected
    if not selected_year_id:
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
        if current_academic_year:
            selected_year_id = str(current_academic_year.id)
    
    if not selected_year_id:
        messages.error(request, "No current academic year configured.")
        return redirect("student_performance_view")
    
    # Get the selected academic year object
    current_academic_year = AcademicYear.objects.filter(id=selected_year_id).first()
    if not current_academic_year:
        messages.error(request, "Selected academic year not found.")
        return redirect("student_performance_view")

    # Initial context for class selection
    context = {
        "grades": unique_grades,
        "academic_years": academic_years,
        "selected_year_id": selected_year_id,
        "school_name": getattr(SchoolSetting.load(), "school_name", None),
        "show_selection": not selected_grade or selected_grade not in unique_grades,
    }

    # If no grade is selected, return with only selection form
    if context["show_selection"]:
        return render(request, "results/assessment_sheet.html", context)

    # Proceed with assessment sheet data if a grade is selected
    academic_classes = AcademicClass.objects.filter(
        Class__name=selected_grade, academic_year=current_academic_year
    ).distinct() if selected_grade else AcademicClass.objects.none()
    
    class_ids = academic_classes.values_list("id", flat=True)

    # Terms and selected term handling
    terms = Term.objects.filter(academic_year=current_academic_year).order_by("start_date")
    selected_term_id = request.GET.get("term_id")
    selected_assessment_type_id = request.GET.get("assessment_type_id")
    # Normalize "None"/empty to real None to avoid invalid ID lookups (e.g., export calls with assessment_type_id=None)
    if selected_assessment_type_id in (None, "", "None", "none", "NULL", "null"):
        selected_assessment_type_id = None

    if not selected_term_id:
        selected_term = Term.objects.filter(
            academic_year=current_academic_year, is_current=True
        ).first()
        selected_term_id = selected_term.id if selected_term else (terms.first().id if terms.exists() else None)

    if not selected_term_id:
        messages.error(request, "No valid term available.")
        return redirect("student_performance_view")

    term_obj = get_object_or_404(Term, id=selected_term_id)
    # Ensure an assessment type is selected by default (remove "All Assessments" mode)
    if not selected_assessment_type_id:
        default_at = AssessmentType.objects.filter(name__iexact="BEGINNING OF TERM").first()
        if not default_at:
            default_at = AssessmentType.objects.order_by('name').first()
        selected_assessment_type_id = str(default_at.id) if default_at else None
    assessment_type = get_object_or_404(AssessmentType, id=selected_assessment_type_id) if selected_assessment_type_id else None

    # Order assessment types
    assessment_order = Case(
        When(name__iexact="BEGINNING OF TERM", then=Value(1)),
        When(name__iexact="MID OF TERM", then=Value(2)),
        When(name__iexact="END OF TERM INTERNAL", then=Value(3)),
        When(name__iexact="END OF TERM EXTERNAL", then=Value(4)),
        default=Value(5),
        output_field=IntegerField(),
    )
    assessment_types = AssessmentType.objects.all().order_by(assessment_order, "name")

    # Registers
    class_registers = ClassRegister.objects.filter(
        academic_class_stream__academic_class__id__in=class_ids,
        academic_class_stream__academic_class__term_id=selected_term_id,
        student__is_active=True,
    ).select_related("student", "academic_class_stream__academic_class").order_by("student__student_name")

    # Dynamically fetch class teacher
    class_teacher = "Tr. [Teacher Name]"
    if academic_classes.exists():
        first_class_stream = AcademicClassStream.objects.filter(
            academic_class__id__in=class_ids,
            academic_class__term_id=selected_term_id,
        ).select_related("class_teacher").first()
        if first_class_stream and first_class_stream.class_teacher:
            class_teacher = f"Tr. {first_class_stream.class_teacher.first_name} {first_class_stream.class_teacher.last_name}"

    # Build index of subject name -> subject ids within scope, then collect teachers
    subject_name_to_ids = defaultdict(set)
    subject_teachers_map = defaultdict(set)
    alloc_map = {}
    if academic_classes.exists():
        related_subjects = Subject.objects.filter(
            assessments__academic_class__id__in=class_ids
        ).distinct()
        for s in related_subjects:
            subject_name_to_ids[norm_key(getattr(s, "name", ""))].add(s.id)
        subject_allocations = ClassSubjectAllocation.objects.filter(
            academic_class_stream__academic_class__id__in=class_ids,
            academic_class_stream__academic_class__term_id=selected_term_id,
        ).select_related("subject_teacher", "subject", "academic_class_stream")
        for allocation in subject_allocations:
            subj_key = norm_key(getattr(allocation.subject, "name", ""))
            teacher_name = f"Tr. {allocation.subject_teacher.first_name} {allocation.subject_teacher.last_name}".strip()
            if teacher_name:
                subject_teachers_map[subj_key].add(teacher_name)
            alloc_map[(allocation.academic_class_stream_id, allocation.subject_id)] = teacher_name

    unique_subjects = []
    seen_subjects = set()
    students_data = []
    # Track subject ids encountered per normalized subject key from results
    subject_ids_by_key = defaultdict(set)

    for register in class_registers:
        student = register.student
        results = Result.objects.filter(
            student=student,
            assessment__academic_class__term_id=selected_term_id,
            student__is_active=True,
        ).select_related("assessment__subject", "assessment__assessment_type")

        if assessment_type:
            results = results.filter(assessment__assessment_type=assessment_type)

        subjects_payload = {}
        total_marks_sum = Decimal("0")
        total_aggregates_sum = 0

        for r in results:
            subj_name = r.assessment.subject.name or ""
            subj_desc = r.assessment.subject.description or ""
            subj_key = norm_key(subj_name)
            subject_ids_by_key[subj_key].add(r.assessment.subject_id)

            if subj_key not in seen_subjects:
                unique_subjects.append(subj_key)
                seen_subjects.add(subj_key)

            # Displayed score uses weighted actual_score
            score = r.actual_score

            # Robust points resolution:
            # 1) Try model property (GradingSystem)
            # 2) Try selector helper get_grade_and_points using RAW score
            # 3) Final fallback to hard-coded scale
            raw_score = r.score
            points = None
            try:
                pval = r.points
                points = float(pval) if pval is not None else None
            except Exception:
                points = None

            if points is None or points == 0.0:
                try:
                    g, p = get_grade_and_points(Decimal(str(raw_score)))
                    points = float(p) if p is not None else None
                except Exception:
                    points = None

            if points is None or points <= 0:
                try:
                    s = float(raw_score)
                    if s >= 80:
                        points = 1.0
                    elif s >= 75:
                        points = 2.0
                    elif s >= 70:
                        points = 3.0
                    elif s >= 65:
                        points = 4.0
                    elif s >= 60:
                        points = 5.0
                    elif s >= 55:
                        points = 6.0
                    elif s >= 50:
                        points = 7.0
                    elif s >= 40:
                        points = 8.0
                    else:
                        points = 9.0
                except Exception:
                    points = 0.0

            subjects_payload[subj_key] = {
                "score": int(score) if score is not None else 0,
                "agg": points,
            }

            # Map subject teacher using stream-specific allocation if available
            teacher_name = alloc_map.get((register.academic_class_stream_id, r.assessment.subject_id))
            if teacher_name:
                subject_teachers_map[subj_key].add(teacher_name)
            else:
                # Stream-specific fuzzy fallback: try matching by subject name/description within this stream
                try:
                    fallback_alloc_stream = (
                        ClassSubjectAllocation.objects.filter(
                            academic_class_stream_id=register.academic_class_stream_id
                        )
                        .filter(
                            Q(subject__name__iexact=subj_name)
                            | Q(subject__name__icontains=subj_name)
                            | Q(subject__description__icontains=subj_name)
                        )
                        .select_related("subject_teacher")
                        .first()
                    )
                    if fallback_alloc_stream:
                        tname2 = f"Tr. {fallback_alloc_stream.subject_teacher.first_name} {fallback_alloc_stream.subject_teacher.last_name}".strip()
                        if tname2:
                            subject_teachers_map[subj_key].add(tname2)
                except Exception:
                    pass

            if not is_excluded_subject(subj_name, subj_desc):
                total_marks_sum += Decimal(str(score)) if score is not None else Decimal("0")
                total_aggregates_sum += float(points) if points is not None else 0.0

        subject_grades = {
            subject: {
                "grade": "",
                "points": data.get("agg")
            }
            for subject, data in subjects_payload.items()
        }

        if total_aggregates_sum:
            division_label, _ = get_division_with_override(
                int(total_aggregates_sum),
                subject_grades
            )
        else:
            division_label = "-"

        student_record = {
            "name": student.student_name,
            "subjects": subjects_payload,
            "total_marks": int(q2(total_marks_sum)),
            "total_aggregates": total_aggregates_sum,
            "division": division_label,
        }

        students_data.append(student_record)

    # Initialize subject_grade_dist with all unique_subjects
    subject_grade_dist = {subject: {"A": 0, "B": 0, "C": 0, "D": 0, "F": 0} for subject in unique_subjects}

    # Calculate grade distribution for each subject
    for student in students_data:
        for subject, data in student['subjects'].items():
            score = data.get('score', 0)
            if score >= 80:
                subject_grade_dist[subject]["A"] += 1
            elif score >= 70:
                subject_grade_dist[subject]["B"] += 1
            elif score >= 60:
                subject_grade_dist[subject]["C"] += 1
            elif score >= 50:
                subject_grade_dist[subject]["D"] += 1
            else:
                subject_grade_dist[subject]["F"] += 1

    # Calculate division counts
    division_counts = {1: 0, 2: 0, 3: 0, 4: 0, "U": 0}
    for student in students_data:
        division = student["division"]
        if division in division_counts:
            division_counts[division] += 1

    # Compute PASS PERCENTAGE (>= 70%) per subject
    subject_pass_percentage = {}
    for subject in unique_subjects:
        total = 0
        passed = 0
        for s in students_data:
            subj = s.get('subjects', {}).get(subject)
            if subj is not None:
                try:
                    score_val = float(subj.get('score', 0))
                except (TypeError, ValueError):
                    score_val = 0
                total += 1
                if score_val >= 70:
                    passed += 1
        subject_pass_percentage[subject] = round((passed / total) * 100) if total else 0

    # Calculate colspan for the empty message
    colspan = len(unique_subjects) * 2 + 4  # 2 columns per subject (score and AGG) + No, Name, Total Marks, Total AGG, Division

    # Flatten subject teachers map to printable dict with robust fallbacks per subject
    subject_teachers = {}
    for subj in unique_subjects:
        names = set(subject_teachers_map.get(subj, set()))
        if not names:
            try:
                # Prefer exact subject ids seen in results for this normalized key
                ids_from_results = list(subject_ids_by_key.get(subj, set()))
                # Also include ids collected by subject name across class assessments
                ids_from_names = list(subject_name_to_ids.get(subj, set()))
                combined_ids = list({*ids_from_results, *ids_from_names})

                # If still empty, try to resolve subject ids by direct lookup (name/description)
                if not combined_ids:
                    direct_ids = list(
                        Subject.objects.filter(name__iexact=subj).values_list("id", flat=True)
                    )
                    if not direct_ids:
                        direct_ids = list(
                            Subject.objects.filter(
                                Q(name__icontains=subj) | Q(description__icontains=subj)
                            ).values_list("id", flat=True)
                        )
                    combined_ids = direct_ids

                # Reusable base queryset
                qs_base = ClassSubjectAllocation.objects.select_related(
                    "subject_teacher", "subject", "academic_class_stream__academic_class"
                )

                def add_names(qs):
                    for alloc in qs:
                        tname = f"Tr. {alloc.subject_teacher.first_name} {alloc.subject_teacher.last_name}".strip()
                        if tname:
                            names.add(tname)

                # 1) Same grade + selected term
                if combined_ids:
                    qs = qs_base.filter(
                        academic_class_stream__academic_class__id__in=class_ids,
                        academic_class_stream__academic_class__term_id=selected_term_id,
                        subject_id__in=combined_ids,
                    )
                else:
                    qs = qs_base.filter(
                        academic_class_stream__academic_class__id__in=class_ids,
                        academic_class_stream__academic_class__term_id=selected_term_id,
                    ).filter(
                        Q(subject__name__iexact=subj)
                        | Q(subject__name__icontains=subj)
                        | Q(subject__description__icontains=subj)
                    )
                add_names(qs)

                # 2) Same grade, any term in current academic year
                if not names:
                    if combined_ids:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__id__in=class_ids,
                            subject_id__in=combined_ids,
                        )
                    else:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__id__in=class_ids,
                        ).filter(
                            Q(subject__name__iexact=subj)
                            | Q(subject__name__icontains=subj)
                            | Q(subject__description__icontains=subj)
                        )
                    add_names(qs)

                # 3) Any grade in current academic year, selected term
                if not names:
                    if combined_ids:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__academic_year=current_academic_year,
                            academic_class_stream__academic_class__term_id=selected_term_id,
                            subject_id__in=combined_ids,
                        )
                    else:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__academic_year=current_academic_year,
                            academic_class_stream__academic_class__term_id=selected_term_id,
                        ).filter(
                            Q(subject__name__iexact=subj)
                            | Q(subject__name__icontains=subj)
                            | Q(subject__description__icontains=subj)
                        )
                    add_names(qs)

                # 4) Any grade in current academic year, any term
                if not names:
                    if combined_ids:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__academic_year=current_academic_year,
                            subject_id__in=combined_ids,
                        )
                    else:
                        qs = qs_base.filter(
                            academic_class_stream__academic_class__academic_year=current_academic_year,
                        ).filter(
                            Q(subject__name__iexact=subj)
                            | Q(subject__name__icontains=subj)
                            | Q(subject__description__icontains=subj)
                        )
                    add_names(qs)

                # 5) Global fallback by subject name anywhere
                if not names:
                    qs = qs_base.filter(
                        Q(subject__name__iexact=subj)
                        | Q(subject__name__icontains=subj)
                        | Q(subject__description__icontains=subj)
                    )
                    add_names(qs)

            except Exception:
                # If anything goes wrong, keep placeholder and continue
                pass

        subject_teachers[subj] = ", ".join(sorted(names)) if names else "Tr. [Teacher Name]"
    # Handle export
    export_fmt = request.GET.get('export')
    if export_fmt == 'csv':
        # Build CSV export of the main table
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = ["NO", "NAME"]
        for subject in unique_subjects:
            headers.extend([subject, "AGG"])
        headers.extend(["T.T MARKS", "T.T AGG", "DIV"])
        writer.writerow(headers)

        # Rows
        for i, student in enumerate(students_data, 1):
            row = [str(i), (student.get('name') or '').title()]
            for subject in unique_subjects:
                subj_data = (student.get('subjects') or {}).get(subject, {}) or {}
                score = subj_data.get('score', '-')
                agg = subj_data.get('agg', '-')
                row.extend([score, agg])
            row.extend([
                student.get('total_marks', '-'),
                student.get('total_aggregates', '-'),
                student.get('division', '-'),
            ])
            writer.writerow(row)

        csv_content = output.getvalue()
        output.close()

        school_name_text = getattr(SchoolSetting.load(), "school_name", None) or "School"
        filename = f'{school_name_text}_{term_obj.term}_{selected_grade}_Assessment.csv'.replace(' ', '_')
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    elif export_fmt == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        normal_style = styles['Normal']

        # Styled header (logo + titles)
        school = None
        school_name_text = request.session.get('school_name', 'Bayan Learning Center')
        try:
            school = SchoolSetting.load()
            if getattr(school, 'school_name', None):
                school_name_text = school.school_name
        except Exception:
            pass

        logo_img = None
        try:
            if school and getattr(school, 'school_logo', None) and getattr(school.school_logo, 'path', None):
                logo_img = Image(school.school_logo.path, width=60, height=60)
        except Exception:
            logo_img = None

        # Header right block
        title_text = f"ASSESSMENT SHEET"
        subtitle_text = f"{(assessment_type.name if assessment_type else '-').upper()} | TERM: {term_obj.term.upper()}"
        class_text = f"CLASS {selected_grade} | CLASSTEACHER: {class_teacher}"

        title_para = Paragraph(f"<para align='center'><b>{school_name_text}</b></para>", styles['Title'])
        subtitle_para = Paragraph(f"<para align='center' spaceb='4'>{title_text}</para>", styles['Heading2'])
        detail_para = Paragraph(f"<para align='center' spaceb='4'>{subtitle_text}</para>", styles['Heading4'])
        class_para = Paragraph(f"<para align='center'>{class_text}</para>", styles['Normal'])

        header_data = [
            [logo_img if logo_img else Paragraph("", normal_style),
             [title_para, subtitle_para, detail_para, class_para]]
        ]
        header_table = Table(header_data, colWidths=[70, 700])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 12))

        # Prepare table with two-tier header to allow long subject names to wrap
        top_header = ["NO", "NAME"]
        for subject in unique_subjects:
            # Use Paragraph with smaller font so long names wrap across both subject columns
            subj_para = Paragraph(f'<para align="center"><b><font size="8">{subject}</font></b></para>', styles["Normal"])
            top_header.extend([subj_para, ""])
        top_header.extend(["T.T MARKS", "T.T AGG", "DIV"])

        # Second header row shows column labels for the two subject columns
        sub_header = ["", ""]
        for _ in unique_subjects:
            sub_header.extend(["MARKS", "AGG"])
        sub_header.extend(["", "", ""])

        table_data = [top_header, sub_header]
        for i, student in enumerate(students_data, 1):
            row = [str(i), student['name'].title()]
            for subject in unique_subjects:
                row.extend([
                    str(student['subjects'].get(subject, {}).get('score', '-')),
                    str(student['subjects'].get(subject, {}).get('agg', '-'))
                ])
            row.extend([str(student['total_marks']), str(student['total_aggregates']), student['division']])
            table_data.append(row)

        # Dynamic, responsive column widths based on longest word in each subject label
        page_w, _ = landscape(letter)
        available_w = page_w - 36 - 36  # left/right margins used in SimpleDocTemplate

        # Fixed columns
        no_w = 30
        name_w = 150
        tt_marks_w = 70
        tt_agg_w = 70
        div_w = 40
        fixed_w = no_w + name_w + tt_marks_w + tt_agg_w + div_w
        subjects_area_w = max(120.0, available_w - fixed_w)

        # Compute pair widths proportional to longest word length; enforce sensible minimums
        pair_raw = []
        for s in unique_subjects:
            text = str(s or "")
            words = [w for w in text.split(" ") if w]
            longest = max((len(w) for w in words), default=len(text))
            est_pair_width = max(80.0, longest * 6.5)  # approx points required at small header font
            pair_raw.append(est_pair_width)

        total_raw = sum(pair_raw) if pair_raw else 0.0
        scale = (subjects_area_w / total_raw) if total_raw and total_raw > subjects_area_w else 1.0
        pair_final = [(w * scale) for w in pair_raw] if pair_raw else []

        # Split each pair into MARKS and AGG columns (MARKS ~60%, AGG ~40%) with hard mins
        dynamic_subject_colwidths = []
        for w in pair_final:
            marks_w = max(35.0, w * 0.60)
            agg_w = max(25.0, w - marks_w)
            dynamic_subject_colwidths.extend([marks_w, agg_w])

        col_widths = [no_w, name_w] + dynamic_subject_colwidths + [tt_marks_w, tt_agg_w, div_w]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Header rows (apply to both the top and second header rows)
            ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 1), colors.white),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 8),
            ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 1), 6),

            # Body
            ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.whitesmoke, colors.HexColor('#f7f7f7')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('FONTNAME', (0, 2), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 2), (-1, -1), 8),
            ('ALIGN', (0, 2), (0, -1), 'CENTER'),          # NO
            ('ALIGN', (1, 2), (1, -1), 'LEFT'),            # NAME
            ('ALIGN', (2, 2), (-1, -1), 'CENTER'),         # Others
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),

            # Repeat both header rows on new pages
            ('REPEATROWS', (0, 0), (-1, 1)),
        ]))

        # Span "NO" and "NAME" vertically across both header rows
        table.setStyle(TableStyle([
            ('SPAN', (0, 0), (0, 1)),
            ('SPAN', (1, 0), (1, 1)),
        ]))

        # Span each subject label across its two columns in the first header row
        subj_count = len(unique_subjects)
        for j in range(subj_count):
            start_col = 2 + (j * 2)
            table.setStyle(TableStyle([('SPAN', (start_col, 0), (start_col + 1, 0))]))

        # Span the final totals columns across both header rows
        last_start = 2 + (subj_count * 2)
        table.setStyle(TableStyle([
            ('SPAN', (last_start + 0, 0), (last_start + 0, 1)),  # T.T MARKS
            ('SPAN', (last_start + 1, 0), (last_start + 1, 1)),  # T.T AGG
            ('SPAN', (last_start + 2, 0), (last_start + 2, 1)),  # DIV
        ]))

        # Add main table
        elements.append(table)

        # Spacer
        elements.append(Spacer(1, 12))

        # ---- Subjects Pass Percentage (>= 70%) ----
        heading_pp = Paragraph("SUBJECTS PASS PERCENTAGE (>=70%)", title_style)
        elements.append(heading_pp)
        elements.append(Spacer(1, 6))

        pp_headers = ["NO", "SUBJECT", "PASS PERCENTAGE(70% AND ABOVE)", "SUBJECT TEACHER"]
        pp_data = [pp_headers]
        for i, subject in enumerate(unique_subjects, 1):
            pass_pct = subject_pass_percentage.get(subject, 0)
            teacher = subject_teachers.get(subject, "Tr. [Teacher Name]")
            pp_data.append([str(i), subject, f"{pass_pct}%", teacher])

        pp_table = Table(pp_data, colWidths=[30, 150, 220, 220])
        pp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(pp_table)

        # Spacer
        elements.append(Spacer(1, 12))

        # ---- Division Counts ----
        heading_div = Paragraph("NO. OF PUPILS AND THEIR DIVISION", title_style)
        elements.append(heading_div)
        elements.append(Spacer(1, 6))

        div_headers = ["DIVISION", "1", "2", "3", "4", "U"]
        div_row = [
            "NUMBER OF PUPILS",
            str(division_counts.get(1, 0)),
            str(division_counts.get(2, 0)),
            str(division_counts.get(3, 0)),
            str(division_counts.get(4, 0)),
            str(division_counts.get("U", 0)),
        ]
        div_data = [div_headers, div_row]

        div_table = Table(div_data, colWidths=[120, 60, 60, 60, 60, 60])
        div_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(div_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        # Create response
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{request.session.get("school_name", "School")}_{term_obj.term}_{selected_grade}_Assessment.pdf"'
        return response

    # Update context with assessment sheet data and grade distribution
    context.update({
        "term": term_obj.term,
        "class_teacher": class_teacher,
        "students_data": students_data,
        "subjects": unique_subjects,
        "terms": terms,
        "selected_term_id": str(selected_term_id),
        "selected_grade": selected_grade,
        "assessment_types": assessment_types,
        "selected_assessment_type_id": selected_assessment_type_id,
        "selected_assessment_type_name": (assessment_type.name if assessment_type else "-"),
        "subject_teachers": subject_teachers,
        "colspan": colspan,
        "show_selection": False,
        "subject_grade_dist": subject_grade_dist,
        "subject_pass_percentage": subject_pass_percentage,
        "division_counts": division_counts,
    })
    return render(request, "results/assessment_sheet.html", context)




# Bulk Result Entry with CSV
@login_required
def bulk_result_entry_view(request, assessment_id=None):
    if assessment_id:
        assessment = get_object_or_404(Assessment, id=assessment_id)
        denial_reason = _mark_entry_denial_reason(request, assessment)
        if denial_reason:
            messages.error(request, denial_reason)
            return redirect('list_assessments', class_id=assessment.academic_class_id)
    else:
        assessment = None

    if request.method == "POST" and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        assessment_id = request.POST.get('assessment_id')
        assessment = get_object_or_404(Assessment, id=assessment_id)
        denial_reason = _mark_entry_denial_reason(request, assessment)
        if denial_reason:
            messages.error(request, denial_reason)
            return redirect('list_assessments', class_id=assessment.academic_class_id)
        batch = ensure_batch_for_assessment(assessment)
        if batch.status != "DRAFT":
            messages.error(request, "Results are locked after submission for verification.")
            return redirect('add_results', assessment_id=assessment.id)

        # Process CSV
        import csv
        import io
        data = csv_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(data))

        created_count = 0
        updated_count = 0
        error_count = 0
        with transaction.atomic():
            for row in reader:
                student_reg_no = (row.get('reg_no') or "").strip()
                score_raw = (row.get('score') or "").strip()
                if not student_reg_no or not score_raw:
                    error_count += 1
                    messages.error(request, "CSV row skipped: reg_no and score are required.")
                    continue
                try:
                    score = _parse_score_or_raise(score_raw, assessment, student_reg_no)
                    student = Student.objects.get(reg_no=student_reg_no, is_active=True)
                    _result_obj, created = Result.objects.update_or_create(
                        assessment=assessment,
                        student=student,
                        defaults={
                            "score": score,
                            "batch": batch,
                            "status": "DRAFT",
                        },
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Student.DoesNotExist:
                    error_count += 1
                    messages.error(request, f"Error processing row for {student_reg_no}: student not found.")
                except ValueError as exc:
                    error_count += 1
                    messages.error(request, f"Error processing row for {student_reg_no}: {exc}")

        if created_count or updated_count:
            messages.success(
                request,
                f"Bulk results processed. Created {created_count}, updated {updated_count}.",
            )
        if error_count and not (created_count or updated_count):
            messages.error(request, "No valid CSV rows were processed.")
        return redirect('add_results', assessment_id=assessment.id)

    # Get all assessments for selection
    assessments = Assessment.objects.all()
    context = {
        'assessment': assessment,
        'assessments': assessments,
    }
    return render(request, 'results/bulk_result_entry.html', context)


@login_required
def verification_queue_view(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    batch = ensure_batch_for_assessment(assessment)
    if not _has_verification_access(request):
        messages.error(request, "You do not have permission to verify results.")
        return redirect('class_assessment_list')

    role_key = _effective_role_key(request)
    samples_qs = VerificationSample.objects.filter(
        result__batch=batch,
        result__student__is_active=True,
    ).select_related(
        "result",
        "result__student",
    ).order_by("result__student__student_name", "result__student__reg_no")
    samples = list(samples_qs)
    sample_map = {sample.result_id: sample for sample in samples}
    sampled_results = [sample.result for sample in samples]

    total_scripts = Result.objects.filter(
        assessment=assessment,
        student__is_active=True,
    ).count()
    sampled_count = len(samples)
    reviewed_count = len([sample for sample in samples if sample.checked_at])
    pending_sample_count = max(sampled_count - reviewed_count, 0)
    sampled_progress_percent = int((reviewed_count / sampled_count) * 100) if sampled_count else 0
    can_finalize_verification = batch.submitted_by_id != request.user.id

    if request.method == "POST":
        if batch.status != "PENDING":
            messages.error(request, "Batch is not pending verification.")
            return redirect('verification_queue', assessment_id=assessment.id)

        if not can_finalize_verification:
            messages.error(
                request,
                "You cannot verify a batch that you submitted. Ask another verifier to complete this step.",
            )
            return redirect('verification_queue', assessment_id=assessment.id)

        if "finalize_verification" in request.POST:
            if sampled_count == 0:
                messages.error(request, "No verification samples are available for this batch.")
                return redirect('verification_queue', assessment_id=assessment.id)

            for sample in samples:
                result = sample.result
                mark_raw = request.POST.get(f"dos_mark_{result.id}")
                if mark_raw not in (None, ""):
                    try:
                        dos_mark = Decimal(mark_raw)
                    except (ValueError, InvalidOperation):
                        messages.error(request, f"Invalid DOS mark for {result.student}.")
                        return redirect('verification_queue', assessment_id=assessment.id)
                    if dos_mark < 0 or dos_mark > Decimal(str(assessment.out_of)):
                        messages.error(
                            request,
                            f"Verifier mark for {result.student} must be between 0 and {assessment.out_of}.",
                        )
                        return redirect('verification_queue', assessment_id=assessment.id)

                    update_sample_mark(sample, dos_mark, request.user)

            remaining_unchecked = VerificationSample.objects.filter(
                result__batch=batch,
                checked_at__isnull=True,
            ).count()
            if remaining_unchecked > 0:
                messages.error(
                    request,
                    f"Complete verification for all sampled scripts before finalizing. Remaining: {remaining_unchecked}.",
                )
                return redirect('verification_queue', assessment_id=assessment.id)

            rejection_reason = request.POST.get("rejection_reason", "").strip()
            has_mismatch = VerificationSample.objects.filter(result__batch=batch, matched=False).exists()
            if has_mismatch and not rejection_reason:
                messages.error(request, "Provide a rejection reason for mismatched samples.")
                return redirect('verification_queue', assessment_id=assessment.id)
            status = evaluate_batch_verification(batch, request.user, rejection_reason=rejection_reason)
            if status == "VERIFIED":
                messages.success(request, "Batch verified successfully.")
            elif status == "FLAGGED":
                messages.error(request, "Mismatches detected. Batch flagged for review.")
            else:
                messages.error(request, "Complete all DOS marks before finalizing.")
            return redirect('verification_queue', assessment_id=assessment.id)

        elif "reject_batch" in request.POST:
            rejection_reason = request.POST.get("rejection_reason", "").strip()
            if not rejection_reason:
                messages.error(request, "Rejection reason is mandatory when explicitly rejecting a batch.")
                return redirect('verification_queue', assessment_id=assessment.id)

            batch.status = "FLAGGED"
            batch.rejection_reason = rejection_reason
            batch.verified_by = request.user
            batch.verified_at = timezone.now()
            batch.save(update_fields=["status", "rejection_reason", "verified_by", "verified_at"])
            Result.objects.filter(batch=batch).update(status="FLAGGED")
            
            # Mark notifications as read
            ResultVerificationNotification.objects.filter(batch=batch, read=False).update(
                read=True,
                read_at=timezone.now(),
            )
            
            _create_verification_report(batch, request.user)
            
            messages.success(request, "Batch rejected and flagged for review.")
            return redirect('verification_queue', assessment_id=assessment.id)

    # Policy: hide teacher marks in the queue for DOS/Admin to avoid bias while re-entering samples.
    show_teacher_mark = False
    context = {
        'assessment': assessment,
        'batch': batch,
        'samples': samples,
        'results': sampled_results,
        'sample_map': sample_map,
        'show_teacher_mark': show_teacher_mark,
        'reviewed_count': reviewed_count,
        'sampled_count': sampled_count,
        'pending_sample_count': pending_sample_count,
        'sampled_progress_percent': sampled_progress_percent,
        'total_scripts': total_scripts,
        'can_finalize_verification': can_finalize_verification,
    }
    return render(request, 'results/verification_queue.html', context)


@login_required
def verification_report_view(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    batch = ensure_batch_for_assessment(assessment)

    if not _has_verification_access(request):
        messages.error(request, "You do not have permission to view verification reports.")
        return redirect('class_assessment_list')

    report = ResultVerificationReport.objects.filter(batch=batch).select_related("verified_by").first()
    samples = VerificationSample.objects.filter(
        result__batch=batch,
        result__student__is_active=True,
    ).select_related(
        "result", "result__student"
    )
    discrepancies = VerificationDiscrepancy.objects.filter(
        batch=batch,
        result__student__is_active=True,
    ).select_related(
        "result", "result__student"
    )
    correction_logs = VerificationCorrectionLog.objects.filter(
        batch=batch,
        result__student__is_active=True,
    ).select_related(
        "result", "result__student", "corrected_by"
    )

    final_results = Result.objects.filter(
        batch=batch,
        student__is_active=True,
    ).select_related("student").order_by("student__student_name")

    teacher_performance = []
    try:
        allocations = ClassSubjectAllocation.objects.filter(
            academic_class_stream__academic_class=assessment.academic_class,
            subject=assessment.subject,
        ).select_related("subject_teacher", "academic_class_stream__stream")
        stream_teacher_map = {}
        for allocation in allocations:
            teacher_name = (
                getattr(allocation.subject_teacher, "staff_name", None)
                or str(allocation.subject_teacher)
            )
            stream_teacher_map[allocation.academic_class_stream.stream_id] = teacher_name

        correction_counts = Counter(correction_logs.values_list("result_id", flat=True))
        teacher_metrics = defaultdict(
            lambda: {
                "teacher": "Unassigned",
                "sampled_count": 0,
                "checked_count": 0,
                "mismatch_count": 0,
                "corrections_count": 0,
                "error_rate": 0,
            }
        )
        for sample in samples:
            teacher_name = stream_teacher_map.get(sample.result.student.stream_id, "Unassigned")
            metric = teacher_metrics[teacher_name]
            metric["teacher"] = teacher_name
            metric["sampled_count"] += 1
            if sample.checked_at:
                metric["checked_count"] += 1
            if sample.matched is False:
                metric["mismatch_count"] += 1
            metric["corrections_count"] += correction_counts.get(sample.result_id, 0)

        teacher_performance = sorted(
            teacher_metrics.values(),
            key=lambda item: (item["teacher"] or ""),
        )
        for metric in teacher_performance:
            checked_count = metric["checked_count"]
            metric["error_rate"] = (
                round((metric["mismatch_count"] / checked_count) * 100, 2)
                if checked_count
                else 0
            )
    except Exception:
        teacher_performance = []

    context = {
        "assessment": assessment,
        "batch": batch,
        "report": report,
        "samples": samples,
        "discrepancies": discrepancies,
        "correction_logs": correction_logs,
        "final_results": final_results,
        "teacher_performance": teacher_performance,
    }
    return render(request, "results/verification_report.html", context)


@login_required
def verification_overview_view(request):
    if not _has_verification_access(request):
        messages.error(request, "You do not have permission to view verification overview.")
        return redirect('class_assessment_list')

    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    selected_year_id = to_int(request.GET.get("year_id"))
    selected_term_id = to_int(request.GET.get("term_id"))
    selected_class_id = to_int(request.GET.get("class_id"))
    selected_subject_id = to_int(request.GET.get("subject_id"))
    selected_assessment_type_id = to_int(request.GET.get("assessment_type_id"))
    selected_status = (request.GET.get("status") or "ALL").upper()

    academic_years = AcademicYear.objects.all().order_by('-id')
    terms = Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date') if selected_year_id else Term.objects.all().order_by('academic_year__id', 'start_date')
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    assessment_types = AssessmentType.objects.all().order_by('name')

    assessments = Assessment.objects.select_related('academic_class', 'assessment_type', 'subject', 'result_batch')

    if selected_year_id:
        assessments = assessments.filter(academic_class__academic_year_id=selected_year_id)
    if selected_term_id:
        assessments = assessments.filter(academic_class__term_id=selected_term_id)
    if selected_class_id:
        assessments = assessments.filter(academic_class__Class_id=selected_class_id)
    if selected_subject_id:
        assessments = assessments.filter(subject_id=selected_subject_id)
    if selected_assessment_type_id:
        assessments = assessments.filter(assessment_type_id=selected_assessment_type_id)

    if selected_status in {"PENDING", "DRAFT", "VERIFIED", "FLAGGED"}:
        assessments = assessments.filter(result_batch__status=selected_status)
    elif selected_status == "ALL":
        assessments = assessments.filter(result_batch__status__in=["PENDING", "DRAFT", "VERIFIED", "FLAGGED"])
    else:
        assessments = assessments.filter(result_batch__status="PENDING")

    pending_count = ResultBatch.objects.filter(status="PENDING").count()
    draft_count = ResultBatch.objects.filter(status="DRAFT").count()

    context = {
        "assessments": assessments.order_by('-date', 'assessment_type__name', 'subject__name'),
        "academic_years": academic_years,
        "terms": terms,
        "classes": classes,
        "subjects": subjects,
        "assessment_types": assessment_types,
        "selected_year_id": str(selected_year_id or ''),
        "selected_term_id": str(selected_term_id or ''),
        "selected_class_id": str(selected_class_id or ''),
        "selected_subject_id": str(selected_subject_id or ''),
        "selected_assessment_type_id": str(selected_assessment_type_id or ''),
        "selected_status": selected_status,
        "pending_count": pending_count,
        "draft_count": draft_count,
    }
    return render(request, "results/verification_overview.html", context)


@login_required
def assessment_verified_sheet_view(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    batch = ensure_batch_for_assessment(assessment)

    if batch.status != "VERIFIED":
        messages.error(request, "Assessment is not verified yet.")
        return redirect('add_results', assessment_id=assessment.id)

    results = Result.objects.filter(
        assessment=assessment,
        student__is_active=True,
    ).select_related(
        "student",
        "assessment__subject",
        "assessment__academic_class__Class",
    )

    def division_from_aggregates(total_aggregates):
        try:
            total = float(total_aggregates)
        except (TypeError, ValueError):
            return "U"
        if 4 <= total <= 12:
            return 1
        if 13 <= total <= 24:
            return 2
        if 25 <= total <= 28:
            return 3
        if 29 <= total <= 32:
            return 4
        return "U"

    class_stream = AcademicClassStream.objects.filter(
        academic_class=assessment.academic_class
    ).select_related("stream", "class_teacher").first()

    subject_teacher = None
    subject_allocation = ClassSubjectAllocation.objects.filter(
        academic_class_stream=class_stream,
        subject=assessment.subject,
    ).select_related("subject_teacher").first() if class_stream else None
    if subject_allocation and subject_allocation.subject_teacher:
        subject_teacher = f"{subject_allocation.subject_teacher.first_name} {subject_allocation.subject_teacher.last_name}".strip()

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = (
        f'attachment; filename="{assessment.subject.name}_{assessment.assessment_type.name}_Verified_Sheet.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=24, bottomMargin=24)
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.HexColor("#1f4e79")
    heading_style = styles['Heading2']
    heading_style.alignment = 1
    heading_style.textColor = colors.HexColor("#1f4e79")

    total_students = results.count()
    academic_year = assessment.academic_class.academic_year.academic_year if assessment.academic_class.academic_year else "-"
    term_label = assessment.academic_class.term.term if assessment.academic_class.term else "-"
    school_name = getattr(SchoolSetting.load(), "school_name", "School")

    header_table = Table([
        [school_name, "Verified Assessment Sheet"],
        [f"{assessment.subject.name} ({assessment.assessment_type.name})", ""],
    ], colWidths=[420, 300])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#e8eef5")),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor("#1f4e79")),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10))

    meta_table = Table([
        ["Academic Year", academic_year, "Term", term_label],
        ["Class", str(assessment.academic_class), "Stream", (class_stream.stream.stream if class_stream else "-")],
        ["Subject Teacher", subject_teacher or "-", "Students", str(total_students)],
        ["Marked Scripts", str(total_students), "Status", batch.status],
    ], colWidths=[110, 210, 90, 170])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f4f6f9")),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#f4f6f9")),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor("#f4f6f9")),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor("#f4f6f9")),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor("#1f2937")),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 12))

    data = [[
        "No",
        "Name",
        assessment.subject.name,
        "AGG",
        "T.T Marks",
        "T.T Agg",
        "DIV",
    ]]

    for idx, result in enumerate(results, 1):
        aggregates = result.points
        division = division_from_aggregates(aggregates)
        data.append([
            str(idx),
            result.student.student_name,
            str(result.score),
            str(result.points),
            str(result.score),
            str(aggregates),
            division,
        ])

    table = Table(data, repeatRows=1, colWidths=[35, 200, 85, 55, 85, 55, 45])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#dbe7f6")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f8fafc"), colors.white]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


@login_required
def class_assessment_combined_view(request):

    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    def _is_grade_two(class_obj):
        try:
            name = (class_obj.Class.name or '').strip().lower()
        except Exception:
            name = ''
        compact = name.replace(' ', '').replace('.', '')
        if compact in {'p2', 'grade2', 'primary2', 'primarytwo'}:
            return True
        if 'grade' in name and '2' in name:
            return True
        if compact.startswith('p') and '2' in compact:
            return True
        return False

    # Read filters
    selected_year_id = to_int(request.GET.get('academic_year_id'))
    selected_term_id = to_int(request.GET.get('term_id'))
    selected_class_id = to_int(request.GET.get('class_id'))
    report_format = request.GET.get('report_format') or 'standard'
    if report_format not in {'standard', 'tahfiz'}:
        report_format = 'standard'

    # assessment_type_ids may come as multiple values or a single comma-separated string
    at_multi = request.GET.getlist('assessment_type_ids')
    if not at_multi:
        raw = request.GET.get('assessment_type_ids')
        if raw:
            at_multi = [p.strip() for p in raw.split(',') if p and p.strip().isdigit()]
    selected_assessment_type_ids = [to_int(x) for x in at_multi if to_int(x)]

    # Defaults for Academic Year and Term
    if not selected_year_id:
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            selected_year_id = current_year.id
    if selected_year_id and not selected_term_id:
        current_term = Term.objects.filter(academic_year_id=selected_year_id, is_current=True).first()
        if current_term:
            selected_term_id = current_term.id

    # Options for dropdowns
    academic_years = AcademicYear.objects.all().order_by('-id')
    terms = Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date') if selected_year_id else Term.objects.none()
    classes = Class.objects.all().order_by('name')
    assessment_types = AssessmentType.objects.all().order_by('name')

    # Data containers
    class_obj = None
    subjects = []
    selected_assessment_types = []
    students_data = []
    ready = False

    # Resolve core scope only when all required filters are present
    if selected_year_id and selected_term_id and selected_class_id and selected_assessment_type_ids:
        class_obj = AcademicClass.objects.filter(
            academic_year_id=selected_year_id,
            term_id=selected_term_id,
            Class_id=selected_class_id
        ).select_related('Class', 'term', 'academic_year').first()

        if not class_obj:
            messages.warning(request, "No Academic Class found for the selected Academic Year, Term and Class.")
        else:
            # Selected assessment types (ordered by custom order for consistent columns)
            assessment_order = Case(
                When(name__iexact="BEGINNING OF TERM", then=Value(1)),
                When(name__iexact="MID OF TERM", then=Value(2)),
                When(name__iexact="END OF TERM INTERNAL", then=Value(3)),
                When(name__iexact="END OF TERM EXTERNAL", then=Value(4)),
                default=Value(5),
                output_field=IntegerField(),
            )
            selected_assessment_types = list(AssessmentType.objects.filter(id__in=selected_assessment_type_ids).order_by(assessment_order, 'name'))
            if not selected_assessment_types:
                messages.warning(request, "Please choose at least one assessment type.")
            else:
                # Map weights for weighted average per subject
                weights = {a.id: (a.weight or 1) for a in selected_assessment_types}
                at_order = [a.id for a in selected_assessment_types]

                # Use ClassRegister to get students who were enrolled in this academic class
                # This ensures promoted students still appear in historical reports
                student_ids = ClassRegister.objects.filter(
                    academic_class_stream__academic_class=class_obj,
                    student__is_active=True,
                ).values_list('student_id', flat=True).distinct()
                
                students = Student.objects.filter(
                    is_active=True,
                    id__in=student_ids,
                ).order_by('student_name')

                # Subjects in scope (those that appear in assessments within this class and selected assessment types)
                _subjects_qs = Subject.objects.filter(
                    assessments__academic_class=class_obj,
                    assessments__assessment_type_id__in=at_order
                ).distinct().order_by('name')
                # Exclude القرآن for combined report
                _subjects_qs = _subjects_qs.exclude(name__iexact='القرآن')
                subjects = list(_subjects_qs.values_list('name', flat=True))

                # Get all results in 1 query
                results_qs = (
                    Result.objects
                    .filter(
                        assessment__academic_class=class_obj,
                        assessment__assessment_type_id__in=at_order,
                        status="VERIFIED",
                        student__is_active=True,
                    )
                    .select_related('student', 'assessment__subject', 'assessment__assessment_type')
                )

                # Build fast lookup: (student_id, subject_name, assessment_type_id) -> score
                data = {}
                for r in results_qs:
                    s_id = r.student_id
                    subj = (r.assessment.subject.name or '').strip()
                    at_id = r.assessment.assessment_type_id
                    if not subj:
                        continue
                    data.setdefault(s_id, {}).setdefault(subj, {})[at_id] = float(r.score)

                # Assemble per-student rows
                for idx, student in enumerate(students, start=1):
                    per_subject = {}
                    for subj in subjects:
                        # Per assessment type scores
                        scores_by_at = {}
                        sum_w = 0.0
                        sum_ws = 0.0
                        for at_id in at_order:
                            score = None
                            if data.get(student.id, {}).get(subj, {}).get(at_id) is not None:
                                score = data[student.id][subj][at_id]
                                w = float(weights.get(at_id, 1))
                                sum_w += w
                                sum_ws += (score * w)
                            scores_by_at[at_id] = score
                        avg = round(sum_ws / sum_w, 2) if sum_w > 0 else None
                        per_subject[subj] = {
                            'scores': scores_by_at,
                            'avg': avg
                        }

                    students_data.append({
                        'no': idx,
                        'student': student,
                        'subjects': per_subject
                    })

                ready = True

    context = {
        # Filter options and selections
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'assessment_types': assessment_types,
        'selected_year_id': str(selected_year_id or ''),
        'selected_term_id': str(selected_term_id or ''),
        'selected_class_id': str(selected_class_id or ''),
        'selected_assessment_type_ids': [str(x) for x in selected_assessment_type_ids],
        'selected_report_format': report_format,

        # Report data
        'class_obj': class_obj,
        'subjects': subjects,
        'selected_assessment_types': selected_assessment_types,
        'students_data': students_data,
        'ready': ready,
    }
    return render(request, 'results/combined_assessments.html', context)

@login_required
def class_assessment_combined_print(request):
    # Handle comment submission
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        term_id = request.POST.get('term_id')
        class_teacher_remark = request.POST.get('class_teacher_remark', '').strip()
        head_teacher_remark = request.POST.get('head_teacher_remark', '').strip()
        
        if student_id and term_id:
            try:
                student = get_object_or_404(Student, id=student_id, is_active=True)
                term = get_object_or_404(Term, id=term_id)
                
                # Get or create the remark
                remark, created = ReportRemark.objects.get_or_create(
                    student=student,
                    term=term,
                    defaults={
                        'class_teacher_remark': class_teacher_remark,
                        'head_teacher_remark': head_teacher_remark,
                        'created_by': request.user
                    }
                )
                
                if not created:
                    # Update existing remark
                    remark.class_teacher_remark = class_teacher_remark
                    remark.head_teacher_remark = head_teacher_remark
                    remark.save()
                
                messages.success(request, f"Comments saved for {student.student_name}")
            except Exception as e:
                messages.error(request, f"Error saving comments: {str(e)}")
            
            # Redirect back to the same page with GET parameters
            return HttpResponseRedirect(request.get_full_path().split('?')[0] + '?' + request.GET.urlencode())
  
    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    def _is_grade_two(class_obj):
        try:
            name = (class_obj.Class.name or '').strip().lower()
        except Exception:
            name = ''
        compact = name.replace(' ', '').replace('.', '')
        if compact in {'p2', 'grade2', 'primary2', 'primarytwo'}:
            return True
        if 'grade' in name and '2' in name:
            return True
        if compact.startswith('p') and '2' in compact:
            return True
        return False

    # Read filters
    academic_year_id = to_int(request.GET.get('academic_year_id'))
    term_id = to_int(request.GET.get('term_id'))
    class_id = to_int(request.GET.get('class_id'))
    report_format = request.GET.get('report_format') or 'standard'
    report_template_map = {
        'standard': 'results/combined_assessments_print.html',
        'tahfiz': 'results/tahfiz/combined_assessments_print.html',
    }
    if report_format not in report_template_map:
        report_format = 'standard'
    report_template = report_template_map[report_format]

    # assessment_type_ids may come as multiple values or a single comma-separated string
    at_multi = request.GET.getlist('assessment_type_ids')
    if not at_multi:
        raw = request.GET.get('assessment_type_ids')
        if raw:
            at_multi = [p.strip() for p in raw.split(',') if p and p.strip().isdigit()]
    selected_assessment_type_ids = [to_int(x) for x in at_multi if to_int(x)]

    # Defaults for Academic Year and Term (use current if not provided)
    if not academic_year_id:
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            academic_year_id = current_year.id
    if academic_year_id and not term_id:
        current_term = Term.objects.filter(academic_year_id=academic_year_id, is_current=True).first()
        if current_term:
            term_id = current_term.id

    # Resolve class scope
    class_obj = None
    if academic_year_id and term_id and class_id:
        class_obj = AcademicClass.objects.filter(
            academic_year_id=academic_year_id,
            term_id=term_id,
            Class_id=class_id
        ).select_related('Class', 'term', 'academic_year').first()

    if not (class_obj and selected_assessment_type_ids):
        messages.error(request, "Please provide Academic Year, Term, Class and at least one Assessment Type.")
        # Reuse the combined selector page for quick navigation
        return redirect('class_assessment_combined')

    # Selected assessment types (ordered for consistent column order)
    assessment_order = Case(
        When(name__iexact="BEGINNING OF TERM", then=Value(1)),
        When(name__iexact="MID OF TERM", then=Value(2)),
        When(name__iexact="END OF TERM INTERNAL", then=Value(3)),
        When(name__iexact="END OF TERM EXTERNAL", then=Value(4)),
        default=Value(5),
        output_field=IntegerField(),
    )
    selected_assessment_types = list(
        AssessmentType.objects.filter(id__in=selected_assessment_type_ids).order_by(assessment_order, 'name')
    )
    if not selected_assessment_types:
        messages.error(request, "No valid assessment types selected.")
        return redirect('class_assessment_combined')

    # Helpers
    at_order = [a.id for a in selected_assessment_types]
    at_name_by_id = {a.id: a.name for a in selected_assessment_types}
    at_weight = {a.id: (a.weight or 1) for a in selected_assessment_types}

    # Load data
    school = SchoolSetting.load()
    # Use ClassRegister to get students who were enrolled in this academic class
    # This ensures promoted students still appear in historical reports
    student_ids = ClassRegister.objects.filter(
        academic_class_stream__academic_class=class_obj,
        student__is_active=True,
    ).values_list('student_id', flat=True).distinct()

    students = Student.objects.filter(is_active=True, id__in=student_ids).order_by('student_name')

    # Get all results in 1 query (within scope)
    results_qs = (
        Result.objects
        .filter(
            assessment__academic_class=class_obj,
            assessment__assessment_type_id__in=at_order,
            status="VERIFIED",
            student__is_active=True,
        )
        .select_related('student', 'assessment__subject', 'assessment__assessment_type')
    )

    # Organize: student_id -> subject_name -> at_id -> {"score": x, "points": y}
    data = {}
    for r in results_qs:
        s_id = r.student_id
        subj = (r.assessment.subject.name or '').strip()
        if not subj:
            continue
        at_id = r.assessment.assessment_type_id
        try:
            score_val = Decimal(str(r.score))
        except Exception:
            score_val = Decimal('0')

        # Prefer model-computed points (uses GradingSystem) for consistency
        try:
            points_val = float(r.points)
        except Exception:
            points_val = None

        if points_val in (None, 0.0):
            # Use selector helper (also backed by GradingSystem)
            _, points_val = get_grade_and_points(score_val)

        if not points_val:
            # Final fallback scale
            s = float(score_val)
            if s >= 80:
                points_val = 1
            elif s >= 75:
                points_val = 2
            elif s >= 70:
                points_val = 3
            elif s >= 65:
                points_val = 4
            elif s >= 60:
                points_val = 5
            elif s >= 55:
                points_val = 6
            elif s >= 50:
                points_val = 7
            elif s >= 40:
                points_val = 8
            else:
                points_val = 9

        data.setdefault(s_id, {}).setdefault(subj, {})[at_id] = {
            'score': float(score_val),
            'points': float(points_val) if points_val is not None else 0.0,
        }

    # Next term info (for header detail)
    next_term = Term.objects.filter(
        academic_year=class_obj.academic_year,
        start_date__gt=class_obj.term.start_date
    ).order_by('start_date').first()
    next_term_start_date = next_term.start_date if next_term else None

    # Signatures (head teacher is common; class teacher signature will be stream-specific)
    head_teacher_signature = Signature.objects.filter(position="HEAD TEACHER").first()

    printed_at = timezone.localtime()
    is_wide_table = len(selected_assessment_types) >= 4

    reports = []
    for student in students:
        # Subject set for this student (limit to those that appear in data)
        subj_map = data.get(student.id, {})
        if not subj_map:
            student_results = results_qs.filter(student_id=student.id)
            for r in student_results:
                subj = (r.assessment.subject.name or '').strip()
                if not subj:
                    continue
                at_id = r.assessment.assessment_type_id
                try:
                    score_val = Decimal(str(r.score))
                except Exception:
                    score_val = Decimal('0')

                try:
                    points_val = float(r.points)
                except Exception:
                    points_val = None

                if points_val in (None, 0.0):
                    _, points_val = get_grade_and_points(score_val)

                if not points_val:
                    s = float(score_val)
                    if s >= 80:
                        points_val = 1
                    elif s >= 75:
                        points_val = 2
                    elif s >= 70:
                        points_val = 3
                    elif s >= 65:
                        points_val = 4
                    elif s >= 60:
                        points_val = 5
                    elif s >= 55:
                        points_val = 6
                    elif s >= 50:
                        points_val = 7
                    elif s >= 40:
                        points_val = 8
                    else:
                        points_val = 9

                subj_map.setdefault(subj, {})[at_id] = {
                    'score': float(score_val),
                    'points': float(points_val) if points_val is not None else 0.0,
                }
        # Exclude القرآن for combined report
        subjects = sorted([n for n in subj_map.keys() if n != 'القرآن'])

        # Totals per assessment type for student (marks and points)
        assessment_totals = {at.name: {'marks': Decimal('0.0'), 'points': Decimal('0.0'), 'count': 0} for at in selected_assessment_types}
        # Build table-like rows (subject lines)
        report_rows = []

        for subject_name in subjects:
            cell_map = {}  # key: at.name -> {"score": -, "points": -}
            sum_w = Decimal('0.0')
            sum_ws = Decimal('0.0')
            # Exclude specific subjects from totals but still display them in the table
            subject_key = (subject_name or '').upper().replace(' ', '')
            excluded_from_totals = ('READING' in subject_key) or ('RELIGIOUSEDUCATION' in subject_key)

            for at_id in at_order:
                at_name = at_name_by_id[at_id]
                rec = subj_map.get(subject_name, {}).get(at_id)
                if rec:
                    score = Decimal(str(rec['score']))
                    pts = Decimal(str(rec['points']))
                    w = Decimal(str(at_weight.get(at_id, 1)))
                    # weighted avg components
                    sum_w += w
                    sum_ws += (score * w)
                    # cell value
                    cell_map[at_name] = {
                        'score': float(score),
                        'points': float(pts),
                    }
                    # totals (skip excluded subjects from totals)
                    if not excluded_from_totals:
                        assessment_totals[at_name]['marks'] += score
                        assessment_totals[at_name]['points'] += pts
                        assessment_totals[at_name]['count'] += 1
                else:
                    # Ensure key exists to simplify template rendering
                    cell_map[at_name] = {
                        'score': '-',
                        'points': '-',
                    }

            average = float((sum_ws / sum_w).quantize(Decimal('0.01'))) if sum_w > 0 else 0.0

            report_rows.append({
                'subject': subject_name,
                'average': average,
                'assessments': cell_map,  # at.name -> {score, points}
            })

        # Assessment divisions from points per assessment type
        assessment_divisions = {}
        for at in selected_assessment_types:
            tot_pts = int(assessment_totals[at.name]['points'])
            count = assessment_totals[at.name]['count']
            assessment_divisions[at.name] = get_division(tot_pts) if count > 0 else "-"

        subject_grades = {}
        for row in report_rows:
            subject_name = row.get('subject')
            points_val = None
            for at in selected_assessment_types:
                data = row.get('assessments', {}).get(at.name, {})
                pts = data.get('points')
                if isinstance(pts, (int, float)):
                    points_val = pts
                    break
                try:
                    pts_num = float(pts)
                    points_val = pts_num
                    break
                except (TypeError, ValueError):
                    continue
            subject_grades[subject_name] = {"grade": "", "points": points_val}

        # Class teacher signature for the student's stream
        class_teacher_signature = None
        try:
            class_stream = AcademicClassStream.objects.filter(
                academic_class=class_obj,
                stream=student.stream
            ).first()
            class_teacher_signature = class_stream.class_teacher_signature if class_stream else None
        except Exception:
            class_teacher_signature = None

        # Avoid duplicating same image between head and class teacher signature
        try:
            if (
                class_teacher_signature
                and head_teacher_signature
                and getattr(head_teacher_signature, 'signature', None)
                and getattr(class_teacher_signature, 'name', None)
                and class_teacher_signature.name == head_teacher_signature.signature.name
            ):
                class_teacher_signature = None
        except Exception:
            pass

        # Get existing remarks for this student and term
        try:
            remark = ReportRemark.objects.get(student=student, term=class_obj.term)
            class_teacher_remark = remark.class_teacher_remark or ''
            head_teacher_remark = remark.head_teacher_remark or ''
        except ReportRemark.DoesNotExist:
            class_teacher_remark = ''
            head_teacher_remark = ''
        
        total_aggregates = sum(
            assessment_totals[at.name]['points']
            for at in selected_assessment_types
            if assessment_totals[at.name]['count'] > 0
        )
        if total_aggregates:
            selected_division, division_override_note = get_division_with_override(
                int(total_aggregates),
                subject_grades
            )
        else:
            selected_division, division_override_note = "-", None

        reports.append({
            'student': student,
            'report_data': report_rows,
            'assessment_types': selected_assessment_types,
            'term': class_obj.term.term,
            'term_id': class_obj.term.id,
            'academic_year': class_obj.academic_year.academic_year,
            'assessment_totals': assessment_totals,
            'assessment_divisions': assessment_divisions,
            'selected_division': selected_division,
            'division_override_note': division_override_note,
            'next_term_start_date': next_term_start_date,
            'head_teacher_signature': head_teacher_signature,
            'class_teacher_signature': class_teacher_signature,
            'class_teacher_remark': class_teacher_remark,
            'head_teacher_remark': head_teacher_remark,
            'report_reference': f"CAR-{class_obj.id}-{class_obj.term.id}-{student.id}",
        })

    context = {
        'school': school,
        'class_obj': class_obj,
        'reports': reports,
        'report_format': report_format,
        'printed_at': printed_at,
        'is_wide_table': is_wide_table,
    }
    return render(request, report_template, context)

@login_required
def results_overview_dashboard(request):
   
    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    selected_year_id = to_int(request.GET.get('academic_year_id'))
    selected_term_id = to_int(request.GET.get('term_id'))
    selected_class_id = to_int(request.GET.get('class_id'))
 
    # Defaults to current AY and current Term
    if not selected_year_id:
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            selected_year_id = current_year.id
    if selected_year_id and not selected_term_id:
        current_term = Term.objects.filter(academic_year_id=selected_year_id, is_current=True).first()
        if current_term:
            selected_term_id = current_term.id
 
    # Scope academic classes for selected year/term (used for cross-joins)
    class_scope = AcademicClass.objects.all()
    if selected_year_id:
        class_scope = class_scope.filter(academic_year_id=selected_year_id)
    if selected_term_id:
        class_scope = class_scope.filter(term_id=selected_term_id)
    if selected_class_id:
        class_scope = class_scope.filter(Class_id=selected_class_id)
 
    # Main results queryset for analytics
    results_qs = Result.objects.filter(student__is_active=True).select_related(
        'student',
        'assessment__subject',
        'assessment__assessment_type',
        'assessment__academic_class__term',
        'assessment__academic_class__academic_year',
        'assessment__academic_class__Class',
    )
    if selected_year_id:
        results_qs = results_qs.filter(assessment__academic_class__academic_year_id=selected_year_id)
    if selected_term_id:
        results_qs = results_qs.filter(assessment__academic_class__term_id=selected_term_id)
    if selected_class_id:
        results_qs = results_qs.filter(assessment__academic_class__Class_id=selected_class_id)

    total_results = results_qs.count()

    # KPIs
    # Students counted via ClassRegister for the scoped classes (avoids counting alumni or outside-year students)
    try:
        class_ids = list(class_scope.values_list('id', flat=True))
        total_students_scoped = Student.objects.filter(
            is_active=True,
            id__in=ClassRegister.objects.filter(
                academic_class_stream__academic_class_id__in=class_ids,
                student__is_active=True,
            ).values_list('student_id', flat=True)
        ).distinct().count() if class_ids else Student.objects.filter(is_active=True).count()
    except Exception:
        total_students_scoped = Student.objects.filter(is_active=True).count()

    average_score_school = results_qs.aggregate(avg=Avg('score'))['avg'] or 0
    pass_rate_school = (results_qs.filter(score__gte=70).count() / total_results * 100) if total_results else 0
    total_assessments = Assessment.objects.filter(academic_class_id__in=class_scope.values('id')).count()
    total_classes = class_scope.values('Class_id').distinct().count()
    total_subjects = Subject.objects.filter(
        assessments__academic_class_id__in=class_scope.values('id')
    ).distinct().count()
    latest_assessment = Assessment.objects.filter(
        academic_class_id__in=class_scope.values('id')
    ).order_by('-date', '-id').first()
    latest_assessment_date = latest_assessment.date if latest_assessment else None

    # Leaderboards and distributions
    top_students = list(
        results_qs.values('student__student_name', 'student__current_class__name')
        .annotate(average=Avg('score'))
        .order_by('-average')[:10]
    )
    bottom_students = list(
        results_qs.values('student__student_name', 'student__current_class__name')
        .annotate(average=Avg('score'))
        .order_by('average')[:10]
    )

    subject_performance_qs = (
        results_qs.values('assessment__subject__name')
        .annotate(avg=Avg('score'), count=Count('id'))
        .order_by('-avg')
    )
    subject_performance = [
        {
            'assessment__subject__name': row['assessment__subject__name'],
            'avg': float(row['avg'] or 0),
            'count': int(row['count'] or 0),
        }
        for row in subject_performance_qs
    ]
    class_league_qs = (
        results_qs.values('assessment__academic_class__Class__name')
        .annotate(avg=Avg('score'), count=Count('id'))
        .order_by('-avg')
    )
    class_league = [
        {
            'assessment__academic_class__Class__name': row['assessment__academic_class__Class__name'],
            'avg': float(row['avg'] or 0),
            'count': int(row['count'] or 0),
        }
        for row in class_league_qs
    ]
    assessment_type_mix_qs = (
        results_qs.values('assessment__assessment_type__name')
        .annotate(avg=Avg('score'), count=Count('id'))
        .order_by('-avg')
    )
    assessment_type_mix = [
        {
            'assessment__assessment_type__name': row['assessment__assessment_type__name'],
            'avg': float(row['avg'] or 0),
            'count': int(row['count'] or 0),
        }
        for row in assessment_type_mix_qs
    ]

    # Top performer per class (highest average per class)
    class_student_avgs = (
        results_qs
        .values('assessment__academic_class__Class__name', 'student__student_name')
        .annotate(avg=Avg('score'))
        .order_by('assessment__academic_class__Class__name', '-avg')
    )
    top_by_class = {}
    for row in class_student_avgs:
        cname = row.get('assessment__academic_class__Class__name') or 'Unknown'
        if cname not in top_by_class:
            top_by_class[cname] = {
                'student': row.get('student__student_name') or '-',
                'average': float(row.get('avg') or 0.0),
            }
    class_top_performers = [
        {'class': cls, 'student': data['student'], 'average': data['average']}
        for cls, data in top_by_class.items()
    ]
 
    # Streams via class registers (accurate for AY/Term scope)
    try:
        stream_stats_qs = ClassRegister.objects.filter(
            academic_class_stream__academic_class_id__in=class_scope.values('id'),
            student__is_active=True,
        ).values('academic_class_stream__stream__stream').annotate(count=Count('id')).order_by('-count')
        stream_comparison = {row['academic_class_stream__stream__stream'] or 'No Stream': row['count'] for row in stream_stats_qs}
    except Exception:
        stream_comparison = {}

    # Gender averages
    gender_raw = results_qs.values('student__gender').annotate(
        avg_score=Avg('score'),
        student_count=Count('student', distinct=True)
    )
    gender_comparison = {}
    for g in gender_raw:
        if g['student__gender'] == 'M':
            gender_comparison['Male'] = {
                'average': float(g['avg_score'] or 0),
                'count': int(g['student_count'] or 0)
            }
        elif g['student__gender'] == 'F':
            gender_comparison['Female'] = {
                'average': float(g['avg_score'] or 0),
                'count': int(g['student_count'] or 0)
            }

    # Trends across terms in the selected year
    performance_trends = []
    if selected_year_id:
        for t in Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date'):
            term_avg = results_qs.filter(assessment__academic_class__term_id=t.id).aggregate(avg=Avg('score'))['avg'] or 0
            performance_trends.append({'term': t.term, 'average': term_avg})

    # Subject difficulty heatmap (same as subject performance but keyed for table)
    subject_heatmap = {row['assessment__subject__name']: row['avg'] for row in subject_performance}

    # Selected labels
    selected_academic_year_name = ''
    selected_term_name = ''
    if selected_year_id:
        try:
            selected_academic_year_name = AcademicYear.objects.get(id=selected_year_id).academic_year
        except AcademicYear.DoesNotExist:
            selected_academic_year_name = ''
    if selected_term_id:
        try:
            selected_term_name = Term.objects.get(id=selected_term_id).term
        except Term.DoesNotExist:
            selected_term_name = ''

    context = {
        # Filters and labels
        'academic_years': AcademicYear.objects.all().order_by('-id'),
        'terms': Term.objects.filter(academic_year_id=selected_year_id).order_by('start_date') if selected_year_id else Term.objects.none(),
        'classes': Class.objects.all().order_by('name'),
        'selected_academic_year': str(selected_year_id or ''),
        'selected_term': str(selected_term_id or ''),
        'selected_class': str(selected_class_id or ''),
        'selected_academic_year_name': selected_academic_year_name,
        'selected_term_name': selected_term_name,
 
        # KPIs
        'kpi': {
            'total_students': total_students_scoped,
            'average_score_school': average_score_school,
            'pass_rate_school': pass_rate_school,
            'total_assessments': total_assessments,
            'total_classes': total_classes,
            'total_subjects': total_subjects,
            'latest_assessment_date': latest_assessment_date,
        },
 
        # Cards and tables
        'top_students': top_students,
        'bottom_students': bottom_students,
        'subject_performance': subject_performance,
        'class_league': class_league,
        'assessment_type_mix': assessment_type_mix,
        'class_top_performers': class_top_performers,
 
        # Charts
        'stream_comparison': stream_comparison,
        'gender_comparison': gender_comparison,
        'performance_trends': performance_trends,
        'subject_heatmap': subject_heatmap,
    }
    return render(request, 'results/school_results_dashboard.html', context)
